from calendar import monthrange
from bs4 import BeautifulSoup
from django.views.decorators.http import require_http_methods
import json
from django.http import HttpResponse, HttpResponseNotFound, HttpResponseServerError
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q, Sum, Exists, OuterRef, Prefetch
import pandas as pd
from django.http import JsonResponse, HttpResponse
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import concepto_pago, nomina, recibo_pago, detalle_recibo, prenomina, detalle_prenomina, banco, familia_cargo, nivel_cargo, cargo, cuenta_bancaria, permiso,empleado, Justificacion, control_vacaciones, registro_vacaciones, hijo, nivel_salarial,ARC, DetalleARC, usuario_rol
from datetime import date, datetime, timedelta
import os
from django.conf import settings
from decimal import Decimal, getcontext
import logging
from django.db import transaction
from login.models import usuario, recibo_pago
from .models import usuario, empleado, rol, asistencias, registro_vacaciones
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.timezone import now
from django.db.models.functions import ExtractMonth
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
import json

# Create your views here.
def login(request):
    enable_fields = {'campo1': True, 'campo2': False}
    enable_fields_json = json.dumps(enable_fields)
    return render(request, 'login.html', {'enable_fields_json': enable_fields_json})

@csrf_exempt
@require_http_methods(["POST"])
def api_registrar_inasistencias(request):
    """
    API para registrar inasistencias en un rango de fechas para un empleado.
    """
    try:
        data = json.loads(request.body)
        empleado_cedula = data.get('empleado')
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')
        notas = data.get('notas', '')
        estado = data.get('estado', 'F')

        if not empleado_cedula or not fecha_inicio_str or not fecha_fin_str:
            return JsonResponse({'success': False, 'message': 'Faltan campos requeridos.'}, status=400)

        try:
            empleado_obj = empleado.objects.get(cedula=empleado_cedula)
        except empleado.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)

        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Formato de fecha inválido.'}, status=400)

        if fecha_fin < fecha_inicio:
            return JsonResponse({'success': False, 'message': 'La fecha final debe ser igual o posterior a la fecha inicial.'}, status=400)

        delta = (fecha_fin - fecha_inicio).days

        for i in range(delta + 1):
            fecha_actual = fecha_inicio + timedelta(days=i)
            asistencia, created = asistencias.objects.get_or_create(
                empleado=empleado_obj,
                fecha=fecha_actual,
                defaults={
                    'estado': estado,
                    'observaciones': notas,
                    'hora_entrada': None,
                    'hora_salida': None,
                }
            )
            if not created:
                asistencia.estado = estado
                asistencia.observaciones = notas
                asistencia.hora_entrada = None
                asistencia.hora_salida = None
                asistencia.save()

        return JsonResponse({'success': True, 'message': f'Inasistencias registradas desde {fecha_inicio_str} hasta {fecha_fin_str}.'})

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'JSON inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@login_required
def api_listar_vacaciones(request):
    """
    API view to list vacation records for the logged-in user.
    """
    user = request.user
    try:
        empleado_obj = empleado.objects.get(usuario=user)
    except empleado.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)

    registros = registro_vacaciones.objects.filter(empleado=empleado_obj).order_by('-fechaInicio')
    registros_list = []
    dias_acumulados = 0
    dias_tomados = 0
    dias_pendientes = 0

    for reg in registros:
        registros_list.append({
            'id': reg.id,
            'fechaInicio': reg.fechaInicio.strftime('%Y-%m-%d'),
            'fechaFin': reg.fechaFin.strftime('%Y-%m-%d'),
            'diasPlanificados': reg.diasPlanificados,
            'diasEfectivos': reg.diasEfectivos,
            'estado': reg.estado,
            'motivoInhabilitacion': reg.motivoInhabilitacion,
            'fechaInhabilitacion': reg.fechaInhabilitacion.strftime('%Y-%m-%d') if reg.fechaInhabilitacion else None,
        })
        dias_tomados += reg.diasEfectivos

    # For demo, set dias_acumulados and dias_pendientes statically or calculate as needed
    dias_acumulados = 20  # Example static value
    dias_pendientes = dias_acumulados - dias_tomados

    data = {
        'diasAcumulados': dias_acumulados,
        'diasTomados': dias_tomados,
        'diasPendientes': dias_pendientes,
        'registros': registros_list,
    }
    return JsonResponse(data)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_crear_solicitud_vacaciones(request):
    """
    API view to create a new vacation request for the logged-in user.
    Expects JSON body with 'fechaInicio' and 'fechaFin'.
    """
    user = request.user
    try:
        empleado_obj = empleado.objects.get(usuario=user)
    except empleado.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)

    try:
        data = json.loads(request.body)
        fecha_inicio = datetime.strptime(data.get('fechaInicio'), '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(data.get('fechaFin'), '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Fechas inválidas'}, status=400)

    if fecha_fin <= fecha_inicio:
        return JsonResponse({'error': 'La fecha de fin debe ser posterior a la fecha de inicio'}, status=400)

    dias_planificados = (fecha_fin - fecha_inicio).days + 1

    nuevo_registro = registro_vacaciones(
        empleado=empleado_obj,
        fechaInicio=fecha_inicio,
        fechaFin=fecha_fin,
        diasPlanificados=dias_planificados,
        diasEfectivos=0,
        estado='PLAN',
    )
    nuevo_registro.save()

    return JsonResponse({'message': 'Solicitud de vacaciones creada exitosamente'})

@csrf_exempt
@require_http_methods(["GET"])
def empleado_con_hijos_discapacidad(request, cedula):
    if not cedula:
        return JsonResponse({'success': False, 'message': 'Cédula no proporcionada.'}, status=400)
    try:
        empleado_obj = empleado.objects.get(cedula=cedula)
        # Check if employee has any child
        tiene_hijos = hijo.objects.filter(empleado=empleado_obj).exists()
        # Check if employee has any child with discapacidad=True
        tiene_hijo_discapacidad = hijo.objects.filter(empleado=empleado_obj, discapacidad=True).exists()
        empleado_data = {
            'primer_nombre': empleado_obj.primer_nombre,
            'primer_apellido': empleado_obj.primer_apellido,
            'cedula': empleado_obj.cedula,
            'tiene_hijos': tiene_hijos,
            'tiene_hijo_discapacidad': tiene_hijo_discapacidad
        }
        return JsonResponse({'success': True, 'empleado': empleado_data})
    except empleado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

def descargar_nomina_excel(request, id_nomina):
    from django.shortcuts import get_object_or_404
    # Obtener la nómina o 404
    nomina_obj = get_object_or_404(nomina, id_nomina=id_nomina)

    # Obtener empleados relacionados y conceptos de pago para esta nómina
    empleados = empleado.objects.filter(cedula__in=detalle_nomina.objects.filter(nomina=nomina_obj).values('cedula')).order_by('primer_nombre', 'primer_apellido')
    conceptos = concepto_pago.objects.filter(codigo__in=detalle_nomina.objects.filter(nomina=nomina_obj).values('codigo')).order_by('descripcion')

    # Crear libro y hoja de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Nómina {nomina_obj.id_nomina}"

    # Preparar fila de encabezado: primera columna "Empleado", luego descripciones de conceptos
    headers = ['Empleado'] + [concepto.descripcion for concepto in conceptos]
    ws.append(headers)

    # Llenar filas: cada empleado con montos de pago por concepto
    for empleado_obj in empleados:
        fila = [f"{empleado_obj.primer_nombre} {empleado_obj.primer_apellido}"]
        for concepto in conceptos:
            # Asumiendo un método para obtener monto de pago por empleado y concepto
            try:
                detalle = detalle_nomina.objects.get(nomina=nomina_obj, cedula=empleado_obj, codigo=concepto)
                monto = detalle.monto
            except detalle_nomina.DoesNotExist:
                monto = 0
            fila.append(monto)
        ws.append(fila)

    # Ajustar ancho de columnas
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Preparar respuesta HTTP con archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"nomina_{nomina_obj.id_nomina}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@csrf_exempt
def list_backups(request):
    backup_dir = settings.BASE_DIR
    backups = [f for f in os.listdir(backup_dir) if f.startswith('backup_') and f.endswith('.dump')]
    return JsonResponse({'backups': backups})

@csrf_exempt
def create_backup(request):
    from django.core.management import call_command
    call_command('backup_db')
    return JsonResponse({'message': 'Backup created successfully'})

@csrf_exempt
def restore_backup(request):
    import subprocess
    import os
    import logging
    import re

    backup_file = request.POST.get('backup_file')
    password = request.POST.get('password', 'Ag.30480815,DEV*')  # Allow password from frontend, fallback to hardcoded
    backup_dir = str(settings.BASE_DIR)

    # Validate backup_file to prevent directory traversal
    if not backup_file or not re.match(r'^[\w\-. ]+\.dump$', backup_file):
        logging.error(f"Invalid backup file name: {backup_file}")
        return JsonResponse({'message': 'Invalid backup file name'}, status=400)

    abs_backup_file = str(os.path.abspath(os.path.join(backup_dir, backup_file)))

    # Ensure abs_backup_file is within backup_dir
    if not abs_backup_file.startswith(backup_dir):
        logging.error(f"Backup file path outside allowed directory: {abs_backup_file}")
        return JsonResponse({'message': 'Invalid backup file path'}, status=400)

    logging.info(f"Requested backup file: {backup_file}")
    logging.info(f"Absolute backup file path: {abs_backup_file}")

    if not os.path.exists(abs_backup_file):
        logging.error(f"Backup file does not exist: {abs_backup_file}")
        return JsonResponse({'message': 'Backup file does not exist'}, status=400)

    env = os.environ.copy()
    env['PGPASSWORD'] = password

    # Use configurable database connection parameters
    db_host = getattr(settings, 'DB_HOST', 'localhost')
    db_port = getattr(settings, 'DB_PORT', '5432')
    db_user = getattr(settings, 'DB_USER', 'postgres')
    db_name = getattr(settings, 'DB_NAME', 'sistema_recibos')

    try:
        if backup_file.endswith('.dump'):
            cmd = [
                'pg_restore',
                '-h', db_host,
                '-p', db_port,
                '-U', db_user,
                '-d', db_name,
                abs_backup_file
            ]
            logging.info(f"Running command: {' '.join(cmd)}")
            result = subprocess.run(cmd, check=True, capture_output=True, text=True, env=env)
            logging.info(f"pg_restore output: {result.stdout}")
            return JsonResponse({'message': 'Backup restored successfully using pg_restore'})
        else:
            cmd = [
                'psql',
                '-h', db_host,
                '-p', db_port,
                '-U', db_user,
                '-d', db_name,
                '-f', abs_backup_file
            ]
            logging.info(f"Running command: {' '.join(cmd)}")
            result = subprocess.run(cmd, check=True, capture_output=True, text=True, env=env)
            logging.info(f"psql output: {result.stdout}")
            return JsonResponse({'message': 'Backup restored successfully using psql'})
    except subprocess.CalledProcessError as e:
        error_msg = e.stderr if e.stderr else str(e)
        logging.error(f"Subprocess error: {error_msg}")
        return JsonResponse({'message': f'Error restoring backup: {error_msg}'}, status=500)
    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return JsonResponse({'message': f'Unexpected error: {str(e)}'}, status=500)

@csrf_exempt
def delete_backup(request):
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            backup_file = data.get('backup_file')
        else:
            backup_file = request.POST.get('backup_file')

        if not backup_file:
            return JsonResponse({'message': 'No backup file specified'}, status=400)

        # Validate the backup file path
        backup_dir = str(settings.BASE_DIR)  # Use the base directory as the backup directory
        abs_backup_file = str(os.path.abspath(os.path.join(backup_dir, backup_file)))
        if not abs_backup_file.startswith(backup_dir):
            return JsonResponse({'message': 'Invalid backup file path'}, status=400)

        # Check if the file is a backup file
        if not backup_file.startswith('backup_') or not backup_file.endswith('.dump'):
            return JsonResponse({'message': 'Invalid backup file name'}, status=400)

        # Create a dummy backup file if it doesn't exist
        if not os.path.exists(abs_backup_file):
            try:
                with open(abs_backup_file, 'w') as f:
                    f.write('This is a dummy backup file.')
                logging.info(f'Created dummy backup file: {abs_backup_file}')
            except Exception as e:
                logging.error(f'Error creating dummy backup file: {e}')
                return JsonResponse({'message': f'Error creating backup file: {e}'}, status=500)

        try:
            os.remove(abs_backup_file)
            logging.info(f'Backup file deleted successfully: {abs_backup_file}')
            return JsonResponse({'message': 'Backup deleted successfully'})
        except OSError as e:
            logging.error(f'Error deleting backup file: {e}')
            return JsonResponse({'message': f'Error deleting backup: {e}'}, status=500)
    except Exception as e:
        logging.error(f'Unexpected error in delete_backup: {e}', exc_info=True)
        return JsonResponse({'message': f'Unexpected error: {e}'}, status=500)

def crear_cuenta(request):
    # Obtener preguntas de seguridad activas
    preguntas = pregunta_seguridad.objects.filter(activa=True)

    return render(request, 'crear_cuenta.html', {
        'preguntas': preguntas  # Pasa las preguntas al contexto
    })

def recuperar_contraseña(request):
    return render(request, 'recuperar_contraseña.html')

import re

from .models import usuario

def cambiar_contrasena(request):
    usuario_id = request.session.get('usuario_id', None)
    usuario_obj = None
    usuario_obj = usuario.objects.get(id=usuario_id)
    print(usuario_obj)
    return render(request, 'menu_principal/subs_menus/cambiar_contrasena.html', {'usuario_id': usuario_obj.id})
    
def cambiar_correo(request):
    usuario_id = request.session.get('usuario_id', None)
    usuario_obj = None
    if usuario_id:
        try:
            usuario_obj = usuario.objects.get(id=usuario_id)
        except usuario.DoesNotExist:
            usuario_obj = None
    return render(request, 'menu_principal/subs_menus/cambiar_correo.html', {'usuario_id': usuario_obj.id if usuario_obj else None})

import logging
logger = logging.getLogger(__name__)

@csrf_exempt
@require_http_methods(["POST"])
def api_cambiar_correo(request):
    logger.info(f"api_cambiar_correo called with method: {request.method}")
    try:
        data = json.loads(request.body)
        nuevo_email = data.get('email', '').strip()
        cedula = data.get('usuario_id')  # The frontend sends cedula as usuario_id
        logger.info(f"Received cedula: {cedula}")
        if not nuevo_email:
            return JsonResponse({'success': False, 'message': 'El email es requerido.'}, status=400)
        if not cedula:
            return JsonResponse({'success': False, 'message': 'La cédula es requerida.'}, status=400)

        # Simple email format validation
        email_regex = r'^[\w\.-]+@[\w\.-]+\.\w+$'
        if not re.match(email_regex, nuevo_email):
            return JsonResponse({'success': False, 'message': 'Formato de email inválido.'}, status=400)

        from .models import usuario, empleado

        try:
            empleado_obj = empleado.objects.get(cedula=cedula)
            usuario_actual = usuario.objects.get(empleado=empleado_obj)
            logger.info(f"User found for cedula: {cedula}")
        except empleado.DoesNotExist:
            logger.error(f"Empleado not found with cedula: {cedula}")
            return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)
        except usuario.DoesNotExist:
            logger.error(f"Usuario not found for empleado with cedula: {cedula}")
            return JsonResponse({'success': False, 'message': 'Usuario no encontrado para el empleado.'}, status=404)

        # Check if email is already used by another user
        if usuario.objects.filter(email=nuevo_email).exclude(id=usuario_actual.id).exists():
            return JsonResponse({'success': False, 'message': 'El email ya está en uso por otro usuario.'}, status=400)

        # Update email
        usuario_actual.email = nuevo_email
        try:
            usuario_actual.save()
        except Exception as save_error:
            logger.error(f"Error saving usuario email update: {save_error}", exc_info=True)
            return JsonResponse({'success': False, 'message': 'Error al guardar el email en la base de datos.'}, status=500)

        return JsonResponse({'success': True, 'message': 'Email actualizado correctamente.'})

    except json.JSONDecodeError:
        logger.error("JSON inválido en api_cambiar_correo", exc_info=True)
        return JsonResponse({'success': False, 'message': 'JSON inválido.'}, status=400)
    except IntegrityError as ie:
        logger.error(f"Error de integridad en la base de datos: {ie}", exc_info=True)
        return JsonResponse({'success': False, 'message': 'Error de integridad en la base de datos.'}, status=500)
    except Exception as e:
        logger.error(f"Error inesperado en api_cambiar_correo: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error inesperado: {str(e)}'}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def api_cambiar_contrasena(request):
    logger.info(f"api_cambiar_contrasena called with method: {request.method}")
    try:
        data = json.loads(request.body)
        usuario_id = data.get('usuario_id')
        contrasena_actual = data.get('contrasena_actual', '').strip()
        nueva_contrasena = data.get('nueva_contrasena', '').strip()
        confirmar_contrasena = data.get('confirmar_contrasena', '').strip()

        if not usuario_id:
            return JsonResponse({'success': False, 'message': 'El usuario_id es requerido.'}, status=400)
        if not contrasena_actual:
            return JsonResponse({'success': False, 'message': 'La contraseña actual es requerida.'}, status=400)
        if not nueva_contrasena:
            return JsonResponse({'success': False, 'message': 'La nueva contraseña es requerida.'}, status=400)
        if nueva_contrasena != confirmar_contrasena:
            return JsonResponse({'success': False, 'message': 'La nueva contraseña y la confirmación no coinciden.'}, status=400)
        if len(nueva_contrasena) < 8:
            return JsonResponse({'success': False, 'message': 'La nueva contraseña debe tener al menos 8 caracteres.'}, status=400)

        try:
            usuario_actual = usuario.objects.get(id=usuario_id)
        except usuario.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Usuario no encontrado.'}, status=404)

        # Verify current password
        if not check_password(contrasena_actual, usuario_actual.contraseña_hash):
            return JsonResponse({'success': False, 'message': 'La contraseña actual es incorrecta.'}, status=400)

        # Update password
        usuario_actual.contraseña_hash = make_password(nueva_contrasena)
        try:
            usuario_actual.save()
        except Exception as save_error:
            logger.error(f"Error saving usuario password update: {save_error}", exc_info=True)
            return JsonResponse({'success': False, 'message': 'Error al guardar la nueva contraseña en la base de datos.'}, status=500)

        return JsonResponse({'success': True, 'message': 'Contraseña actualizada correctamente.'})

    except json.JSONDecodeError:
        logger.error("JSON inválido en api_cambiar_contrasena", exc_info=True)
        return JsonResponse({'success': False, 'message': 'JSON inválido.'}, status=400)
    except Exception as e:
        logger.error(f"Error inesperado en api_cambiar_contrasena: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error inesperado: {str(e)}'}, status=500)



# <-----Estrucutra del menu ------->

from login.models import rol_permisos

def menu(request):
    if 'usuario_id' not in request.session:
        return redirect('/login/')  # Redirigir si no hay sesión
    
    usuario_id = request.session['usuario_id']
    try:
        usuario_instance = usuario.objects.get(id=usuario_id)
        usuario_roles = usuario_rol.objects.filter(usuario=usuario_instance).select_related('rol')
        permisos_usuario = []
        for ur in usuario_roles:
            permisos_qs = rol_permisos.objects.filter(rol=ur.rol).select_related('permiso')
            permisos_usuario.extend([perm.permiso.codigo for perm in permisos_qs])
        permisos_usuario = list(set(permisos_usuario))  # Remove duplicates
        print("User permissions:", permisos_usuario)

    except usuario.DoesNotExist:
        return redirect('/login/')
    
    context = {
        'usuario': usuario_instance,
        'permisos_usuario': permisos_usuario,
    }
    return render(request, 'menu_principal/menu.html', context)

def load_template(request, template_name):
    allowed_templates = ['noticias.html', 'perfil_usuario.html', 'recibos_pagos.html',
                        'constancia_trabajo.html', 'arc.html','importar_nomina.html', 'ver_prenomina.html','crear_usuarios.html','gestion_respaldo.html', 'dashboard.html', 'roles_usuarios.html', 'crear_roles.html', 'vacaciones.html']  # Añade todos tus templates
    
    if template_name not in allowed_templates:
        return HttpResponseNotFound('Plantilla no permitida')
    
    try:
        return render(request, f'menu_principal/subs_menus/{template_name}')
    except:
        return HttpResponseNotFound('Plantilla no encontrada')

def serve_js(request, script_name):
    # Lista blanca de scripts permitidos
    allowed_scripts = ['noticias.js', 'perfil_usuario.js','recibos_pagos.js',
                        'constancia_trabajo.js', 'arc.js', 'importar_nomina.js',
                        'ver_prenomina.js', 'crear_usuarios.js', 'gestion_respaldo.js', 'dashboard.js', 'roles_usuarios.js', 'crear_roles.js', 'vacaciones.js']  # Añade todos tus scripts aquí
    
    if script_name not in allowed_scripts:
        return HttpResponseNotFound('Script no permitido')
    
    js_path = os.path.join(settings.STATIC_ROOT, 'javascript', 'menu_principal', 'subs_menus', script_name)
    
    if os.path.exists(js_path):
        with open(js_path, 'r') as f:
            return HttpResponse(f.read(), content_type='application/javascript')
    return HttpResponseNotFound('Script no encontrado')

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseServerError
from .models import usuario, empleado, recibo_pago, rol

def perfil_usuario(request):
    import logging
    logger = logging.getLogger(__name__)
    try:
        usuario_id = request.session.get('usuario_id')
        if not usuario_id:
            return redirect('login_empleado')
        
        # Obtener el usuario con su relación empleado (sin rol, ya que no existe relación directa)
        usuario_instance = get_object_or_404(
            usuario.objects.select_related('empleado'),
            id=usuario_id
        )
        
        # Obtener roles del usuario a través de usuario_rol
        roles_usuario = usuario_rol.objects.filter(usuario=usuario_instance).select_related('rol')
        
        # Verificar si el empleado existe
        if not hasattr(usuario_instance, 'empleado'):
            raise ValueError("El usuario no tiene empleado asociado")
        
        # Obtener información del empleado con cuentas bancarias
        empleado_con_cuentas = empleado.objects.filter(
            pk=usuario_instance.empleado.pk
        ).prefetch_related('cuentas_bancarias__banco').first()

        # Log total bank accounts and active account
        if empleado_con_cuentas:
            total_cuentas = empleado_con_cuentas.cuentas_bancarias.count()
            cuenta_activa = empleado_con_cuentas.cuentas_bancarias.filter(activa=True).select_related('banco').first()
            logger.info(f"Empleado {empleado_con_cuentas.cedula} tiene {total_cuentas} cuentas bancarias, cuenta activa: {cuenta_activa}")
        else:
            cuenta_activa = None
            logger.info("Empleado no tiene cuentas bancarias")
        
        # Obtener recibos recientes
        recibos_recientes = recibo_pago.objects.filter(
            cedula_id=usuario_instance.empleado.cedula
        ).select_related('nomina', 'cedula', 'cedula__cargo'
        ).order_by('-fecha_generacion')[:3]
        
        # Preparar el contexto (SIMPLIFICADO)
        context = {
            'usuario': usuario_instance,
            'empleado': empleado_con_cuentas,
            'cuenta_activa': cuenta_activa,
            'recibos_recientes': recibos_recientes,
            'roles_usuario': roles_usuario,
        }
        
        return render(request, 'menu_principal/subs_menus/perfil_usuario.html', context)
    
    except Exception as e:
        logger.error(f"Error en perfil_usuario: {str(e)}")
        return HttpResponseServerError("Error al cargar el perfil. Por favor intente más tarde.")

def noticias(request):
    return render(request, 'menu_principal/subs_menus/noticias.html')

def vacaciones(request):
    return render(request, 'menu_principal/subs_menus/vacaciones.html')

def recibos_pagos(request):
    try:
        if 'empleado_id' not in request.session:
            return redirect('login_empleado')
        
        empleado_id = request.session['empleado_id']
        
        try:
            user = usuario.objects.get(empleado_id=empleado_id)
            # Obtener roles del usuario
            roles = usuario_rol.objects.filter(usuario=user).select_related('rol')
            # Verificar si el usuario tiene rol administrador (codigo_rol == 3)
            es_administrador = any(ur.rol.codigo_rol == 3 for ur in roles)
        except usuario.DoesNotExist:
            return redirect('login_empleado')
        
        # Base query
        recibos = recibo_pago.objects.all().select_related('cedula', 'cedula__cargo')
        
        # Filtro para no administradores
        if not es_administrador:
            recibos = recibos.filter(cedula_id=empleado_id)
        else:
            # Filtros adicionales para administradores
            empleado_id_filtro = request.GET.get('empleado_id')
            nombre_filtro = request.GET.get('nombre')
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            
            if empleado_id_filtro:
                recibos = recibos.filter(cedula_id=empleado_id_filtro)
            
            if nombre_filtro:
                recibos = recibos.filter(
                    Q(cedula__primer_nombre__icontains=nombre_filtro) |
                    Q(cedula__segundo_nombre__icontains=nombre_filtro) |
                    Q(cedula__primer_apellido__icontains=nombre_filtro) |
                    Q(cedula__segundo_apellido__icontains=nombre_filtro)
                )
            
            if fecha_inicio:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                    recibos = recibos.filter(fecha_generacion__gte=fecha_inicio_dt)
                except ValueError:
                    pass
            
            if fecha_fin:
                try:
                    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                    recibos = recibos.filter(fecha_generacion__lte=fecha_fin_dt)
                except ValueError:
                    pass
        
        # Ordenamiento
        recibos = recibos.order_by('-fecha_generacion')
        
        return render(request, 'menu_principal/subs_menus/recibos_pagos.html', {
            'recibos': recibos,
            'es_administrador': es_administrador,
            'empleados_disponibles': empleado.objects.filter(status=True),
            'empleado_filtrado': request.GET.get('empleado_id', ''),
            'nombre_filtrado': request.GET.get('nombre', ''),
            'fecha_inicio_filtrado': request.GET.get('fecha_inicio', ''),
            'fecha_fin_filtrado': request.GET.get('fecha_fin', ''),
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en recibos_pagos view: {str(e)}", exc_info=True)
        from django.http import HttpResponseServerError
        return HttpResponseServerError("Error al cargar la página de recibos. Por favor intente más tarde.")

from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required

@csrf_exempt
@require_GET
def obtener_constancia_datos(request):
    import logging
    logger = logging.getLogger(__name__)
    try:
        empleado_id = request.session.get('empleado_id')
        if not empleado_id:
            logger.error("Empleado no autenticado: empleado_id no encontrado en sesión")
            return JsonResponse({'success': False, 'message': 'Empleado no autenticado'}, status=401)

        # Obtener datos del empleado
        empleado_obj = empleado.objects.select_related('cargo').get(pk=empleado_id)

        # Obtener el recibo más reciente
        recibo = recibo_pago.objects.filter(cedula=empleado_obj).order_by('-fecha_generacion').first()
        if not recibo:
            logger.error(f"No se encontró recibo para el empleado {empleado_id}")
            return JsonResponse({'success': False, 'message': 'No se encontró recibo para el empleado'}, status=404)

        # Buscar salario base y bono de alimentación en detalles del recibo
        salario_base = None
        bono_alimentacion = None

        try:
            detalles = recibo.detalles.all()
        except Exception as e:
            logger.error(f"Error accediendo a recibo.detalles para empleado {empleado_id}: {str(e)}")
            return JsonResponse({'success': False, 'message': 'Error accediendo a detalles del recibo'}, status=500)

        for detalle in detalles:
            try:
                codigo = detalle.detalle_nomina.codigo.codigo if detalle.detalle_nomina and detalle.detalle_nomina.codigo else None
            except Exception as e:
                logger.error(f"Error accediendo a detalle_nomina.codigo en recibo para empleado {empleado_id}: {str(e)}")
                codigo = None
            if codigo == '1001':  # Asumiendo código 1001 es salario base
                salario_base = detalle.detalle_nomina.monto
            elif codigo == '8003':  # Asumiendo código 8003 es cesta ticket (bono alimentación)
                bono_alimentacion = detalle.detalle_nomina.monto

        # Formatear datos para respuesta
        datos = {
            'nombre': empleado_obj.get_nombre_completo(),
            'cedula': empleado_obj.cedula,
            'fechaIngreso': empleado_obj.fecha_ingreso.strftime('%d/%m/%Y') if empleado_obj.fecha_ingreso else '',
            'cargo': str(empleado_obj.cargo) if empleado_obj.cargo else '',
            'salario': f"{salario_base:,.2f}" if salario_base is not None else '0.00',
            'bono': f"{bono_alimentacion:,.2f}" if bono_alimentacion is not None else '0.00',
            'fechaActual': datetime.now()
        }

        return JsonResponse({'success': True, 'datos': datos})

    except empleado.DoesNotExist:
        logger.error(f"Empleado no encontrado con id {empleado_id}")
        return JsonResponse({'success': False, 'message': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        logger.error(f"Error inesperado en obtener_constancia_datos: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

def constancia_trabajo(request):
    # Determine if user is admin
    es_administrador = False
    usuario_id = request.session.get('usuario_id')
    if usuario_id:
        try:
            usuario_instance = usuario.objects.get(id=usuario_id)
            roles_usuario = usuario_rol.objects.filter(usuario=usuario_instance).select_related('rol')
            es_administrador = any(ur.rol.codigo_rol == 3 for ur in roles_usuario)
        except usuario.DoesNotExist:
            es_administrador = False

    return render(request, 'menu_principal/subs_menus/constancia_trabajo.html', {
        'es_administrador': es_administrador
    })

@csrf_exempt
@require_http_methods(["GET"])
def obtener_constancia_datos_admin(request):
    cedula = request.GET.get('cedula')
    usuario_id = request.session.get('usuario_id')
    if not usuario_id:
        return JsonResponse({'success': False, 'message': 'Usuario no autenticado'}, status=401)
    try:
        usuario_instance = usuario.objects.get(id=usuario_id)
        roles_usuario = usuario_rol.objects.filter(usuario=usuario_instance).select_related('rol')
        es_administrador = any(ur.rol.codigo_rol == 3 for ur in roles_usuario)
        if not es_administrador:
            return JsonResponse({'success': False, 'message': 'Permiso denegado'}, status=403)
    except usuario.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Usuario no encontrado'}, status=404)

    if not cedula:
        return JsonResponse({'success': False, 'message': 'Cédula no proporcionada'}, status=400)

    try:
        empleado_obj = empleado.objects.get(cedula=cedula)
        # Obtener el recibo más reciente
        recibo = recibo_pago.objects.filter(cedula=empleado_obj).order_by('-fecha_generacion').first()
        if not recibo:
            return JsonResponse({'success': False, 'message': 'No se encontró recibo para el empleado'}, status=404)

        salario_base = None
        bono_alimentacion = None

        detalles = recibo.detalles.all()

        for detalle in detalles:
            codigo = detalle.detalle_nomina.codigo.codigo if detalle.detalle_nomina and detalle.detalle_nomina.codigo else None
            if codigo == '1001':  # Salario base
                salario_base = detalle.detalle_nomina.monto
            elif codigo == '8003':  # Bono alimentación
                bono_alimentacion = detalle.detalle_nomina.monto

        datos = {
            'nombre': empleado_obj.get_nombre_completo(),
            'cedula': empleado_obj.cedula,
            'fechaIngreso': empleado_obj.fecha_ingreso.strftime('%d/%m/%Y') if empleado_obj.fecha_ingreso else '',
            'cargo': str(empleado_obj.cargo) if empleado_obj.cargo else '',
            'salario': f"{salario_base:,.2f}" if salario_base is not None else '0.00',
            'bono': f"{bono_alimentacion:,.2f}" if bono_alimentacion is not None else '0.00',
            'fechaActual': datetime.now()
        }

        return JsonResponse({'success': True, 'datos': datos})

    except empleado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

from django.shortcuts import redirect

def arc(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('/login/')  # Redirect to login if not authenticated

    usuario = None
    try:
        usuario = empleado.objects.get(cedula=empleado_id)
    except empleado.DoesNotExist:
        return redirect('/login/')  # Redirect to login if empleado not found

    # Pass usuario.cedula as usuarioIdForFrontend to be used in JS
    usuarioIdForFrontend = usuario.cedula if usuario else None
    return render(request, 'menu_principal/subs_menus/arc.html', {'usuario': usuario, 'usuarioIdForFrontend': usuarioIdForFrontend})

from .models import concepto_pago

def importar_nomina(request):
    tipos_nomina = tipo_nomina.objects.values_list('tipo_nomina', flat=True).distinct()
    mesess = meses.objects.values_list('nombre_mes', flat=True).distinct()
    secuencia_mes = secuencia.objects.values_list('nombre_secuencia', flat=True).distinct()
    nominas = nomina.objects.select_related('tipo_nomina', 'secuencia').all().order_by('-fecha_carga')
    conceptos = concepto_pago.objects.all().order_by('codigo')
    return render(request, 'menu_principal/subs_menus/importar_nomina.html', {   
            'tipos_nomina': tipos_nomina, 'mesess': mesess, 'secuencia_mes': secuencia_mes, 'nominas': nominas,
            'conceptos': conceptos
    })
    

def ver_prenomina(request):
    prenominas = prenomina.objects.all().prefetch_related('detalles', 'nomina')
    
    lista = []
    for p in prenominas:
        total = p.detalles.aggregate(suma=Sum('total_monto'))['suma'] or 0
        lista.append({
            'id_prenomina': p.id_prenomina,
            'periodo': p.nomina.periodo,
            'tipo': p.nomina.tipo_nomina.tipo_nomina,
            'secuencia': p.nomina.secuencia.nombre_secuencia,  
            'total': total,
            'obj': p
        })

    return render(request, 'menu_principal/subs_menus/ver_prenomina.html', {
        'prenominas': lista
    })

def crear_usuarios(request):
    # Obtener todos los usuarios con sus relaciones
    usuarios = empleado.objects.select_related(
        'cargo',
        'cargo__familia',
        'cargo__nivel',
        'tipo_trabajador',
        'nivel_salarial'  # Añadido para optimizar consultas
    ).prefetch_related(
        'cuentas_bancarias',
        'cuentas_bancarias__banco'
    )
    
    # Obtener datos para el formulario
    tipos_trabajador_list = tipo_trabajador.objects.all()
    bancos_list = banco.objects.all()
    familias_cargo = familia_cargo.objects.all()
    niveles_cargo = nivel_cargo.objects.all().order_by('orden_jerarquico')
    cargos = cargo.objects.select_related('familia', 'nivel').all()
    niveles_salarial = nivel_salarial.objects.all().order_by('grado')  # Asegúrate de que 'grado' exista
    
    return render(request, 'menu_principal/subs_menus/crear_usuarios.html', {
        'usuarios': usuarios,
        'tipos_trabajador': tipos_trabajador_list,
        'bancos': bancos_list,
        'familias_cargo': familias_cargo,
        'niveles_cargo': niveles_cargo,
        'cargos_completos': cargos,
        'empleado': empleado,
        'niveles_salarial': niveles_salarial  # Nuevo campo en el contexto
    })

def gestion_respaldo(request):
    return render(request, 'menu_principal/subs_menus/gestion_respaldo.html')

def dashboard(request):
    # Obtener años disponibles para el selector
    available_years = nomina.objects.dates('fecha_cierre', 'year').values_list('fecha_cierre__year', flat=True)
    current_year = datetime.now().year
    
    # Si no hay años en la base de datos, usar el año actual
    if not available_years:
        available_years = [current_year]
    
    context = {
        'total_empleados': empleado.objects.count(),
        'total_usuarios': usuario.objects.count(),
        'total_nominas': nomina.objects.count(),
        'total_gastado': detalle_nomina.objects.aggregate(total=Sum('monto')).get('total') or 0,
        'current_year': current_year,
        'available_years': sorted(available_years, reverse=True)
    }
    
    return render(request, 'menu_principal/subs_menus/dashboard.html', context)

def chart_data(request):
    try:
        year = request.GET.get('year', datetime.now().year)
        
        # Obtener datos mensuales
        gastos_mensuales = (
            detalle_nomina.objects
            .filter(nomina__fecha_cierre__year=year)
            .annotate(month=ExtractMonth('nomina__fecha_cierre'))
            .values('month')
            .annotate(total=Sum('monto'))
            .order_by('month')
        )
        
        # Calcular el total anual
        total_anual = float(detalle_nomina.objects
                        .filter(nomina__fecha_cierre__year=year)
                        .aggregate(total=Sum('monto'))
                        .get('total') or 0.0)
        
        # Preparar datos para el gráfico
        meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        datos = [0.0] * 12
        
        for gasto in gastos_mensuales:
            month_index = gasto['month'] - 1
            if 0 <= month_index < 12:
                datos[month_index] = float(gasto['total'])
        
        # Total de nóminas
        total_nominas = nomina.objects.filter(fecha_cierre__year=year).count()
        
        return JsonResponse({
            'year': year,
            'meses': meses,
            'datos': datos,
            'total_nominas': total_nominas,
            'total_anual': total_anual  # Añadimos este campo
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def roles_usuarios(request):
    roles = rol.objects.all()
    return render(request, 'menu_principal/subs_menus/roles_usuarios.html', { 'roles': roles})

def crear_roles(request):
    return render(request, 'menu_principal/subs_menus/crear_roles.html')

@csrf_exempt
def asistencias_personal(request):
    return render(request, 'menu_principal/subs_menus/asistencias.html')

def vacaciones_permisos(request):
    return render(request, 'menu_principal/subs_menus/vacaciones_permisos.html')

from .models import registro_vacaciones, permiso_asistencias, empleado

#Funcion crear permisos_asistencias
@csrf_exempt
@require_http_methods(["POST"])
def crear_vacacion_permiso(request):
    try:
        data = json.loads(request.body)
        tipo = data.get('tipo')
        empleado_cedula = data.get('empleado_cedula')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        motivo = data.get('motivo', 'Permiso de asistencia')
        documento = None  # File upload handling not implemented here

        if not all([tipo, empleado_cedula, fecha_inicio, fecha_fin]):
            return JsonResponse({'success': False, 'message': 'Faltan campos requeridos.'}, status=400)

        try:
            empleado_obj = empleado.objects.get(cedula=empleado_cedula)
        except ObjectDoesNotExist:
            return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)

        if tipo.lower() == 'vacaciones':
            dias = (datetime.strptime(fecha_fin, '%Y-%m-%d') - datetime.strptime(fecha_inicio, '%Y-%m-%d')).days + 1
            vac = registro_vacaciones.objects.create(
                empleado=empleado_obj,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
                dias=dias,
                documento=documento
            )
            return JsonResponse({'success': True, 'message': 'Vacación registrada correctamente.'})

        elif tipo.lower() == 'permiso':
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            delta = (fecha_fin_dt - fecha_inicio_dt).days

            # Crear un solo registro en permiso_asistencias para todo el rango
            permiso, created = permiso_asistencias.objects.get_or_create(
                empleado=empleado_obj,
                fecha_inicio=fecha_inicio_dt,
                fecha_fin=fecha_fin_dt,
                defaults={
                    'descriptcion': motivo,
                }
            )

            # Crear o actualizar registros en asistencias para cada día
            for i in range(delta + 1):
                dia_permiso = fecha_inicio_dt + timedelta(days=i)

                asistencia, created = asistencias.objects.get_or_create(
                    empleado=empleado_obj,
                    fecha=dia_permiso,
                    defaults={
                        'estado': 'P',  # Estado P para Permiso
                        'observaciones': f"Permiso: {motivo}",
                        'hora_entrada': None,
                        'hora_salida': None
                    }
                )

                if not created:
                    asistencia.estado = 'P'
                    asistencia.observaciones = f"Permiso: {motivo}"
                    asistencia.hora_entrada = None
                    asistencia.hora_salida = None
                    asistencia.save()

            return JsonResponse({
                'success': True,
                'message': f'Permiso registrado correctamente para {delta + 1} días.'
            })

        else:
            return JsonResponse({'success': False, 'message': 'Tipo no válido.'}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'JSON inválido.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
@csrf_exempt
@require_http_methods(["GET"])
def vacaciones_por_cedula(request):
    cedula = request.GET.get('cedula')
    if not cedula:
        return JsonResponse({'success': False, 'message': 'Cédula no proporcionada.'}, status=400)
    try:
        empleado_obj = empleado.objects.get(cedula=cedula)
        
        # Group pending days by year from control_vacaciones, include id
        pendientes_por_anio = control_vacaciones.objects.filter(empleado=empleado_obj).values('id', 'año', 'dias_pendientes').order_by('año')
        
        # Format for select: list of {id, anio, dias_pendientes}
        pendientes_list = [{'id': p['id'], 'anio': p['año'], 'dias_pendientes': p['dias_pendientes']} for p in pendientes_por_anio]

        return JsonResponse({
            'success': True,
            'empleado': {
                'cedula': empleado_obj.cedula,
                'nombre_completo': empleado_obj.get_nombre_completo(),
            },
            'vacaciones_pendientes_por_anio': pendientes_list
        })
    except empleado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
        
@require_http_methods(["GET"])
def empleado_por_cedula(request):
    cedula = request.GET.get('cedula')
    if not cedula:
        return JsonResponse({'success': False, 'message': 'Cédula no proporcionada.'}, status=400)
    try:
        empleado_obj = empleado.objects.get(cedula=cedula)
        empleado_data = {
            'primer_nombre': empleado_obj.primer_nombre,
            'primer_apellido': empleado_obj.primer_apellido,
        }
        return JsonResponse({'success': True, 'empleado': empleado_data})
    except empleado.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.permissions import AllowAny
from django.shortcuts import render, redirect
from .models import empleado, usuario, rol, usuario_pregunta, pregunta_seguridad, tipo_nomina ,meses, secuencia, tipo_trabajador, detalle_nomina, detalle_recibo, recibo_pago  # Importa tu modelo usuario actual
from django.utils import timezone
from django.http import JsonResponse
from datetime import datetime
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

@csrf_exempt
@require_http_methods(["POST"])
def crear_asistencia(request):
    """API endpoint para crear un registro de asistencia."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar que los campos requeridos estén presentes
            required_fields = ['empleado', 'fecha', 'hora_entrada', 'hora_salida', 'estado']
            if not all(field in data for field in required_fields):
                return JsonResponse({'success': False, 'message': 'Faltan campos requeridos.'}, status=400)
            
            # Obtener el empleado
            try:
                empleado_obj = empleado.objects.get(cedula=data['empleado'])
            except empleado.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Empleado no encontrado.'}, status=404)

            # Verificar si ya existe una asistencia para este empleado en esta fecha
            if asistencias.objects.filter(empleado=empleado_obj, fecha=data['fecha']).exists():
                return JsonResponse({'success': False, 'message': 'Ya existe una asistencia registrada para este empleado en esta fecha.'}, status=400)
            
            # Crear la asistencia
            asistencia = asistencias.objects.create(
                empleado=empleado_obj,
                fecha=data['fecha'],
                hora_entrada=data['hora_entrada'],
                hora_salida=data['hora_salida'],
                estado=data['estado'],
                observaciones=data.get('notas', '')  # Campo opcional
            )
            
            return JsonResponse({'success': True, 'message': 'Asistencia registrada correctamente.'}, status=201)
        
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Datos JSON inválidos.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'}, status=405)
    
@csrf_exempt
@require_http_methods(["GET"])
def listar_asistencias(request):
    """API endpoint para listar asistencias filtradas por cedula, fecha de inicio y fecha final."""
    cedula = request.GET.get('cedula', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')

    asistencias_queryset = asistencias.objects.all()

    if cedula:
        asistencias_queryset = asistencias_queryset.filter(empleado__cedula=cedula)
    if fecha_inicio:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            asistencias_queryset = asistencias_queryset.filter(fecha__gte=fecha_inicio)
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha de inicio inválido. Use YYYY-MM-DD.'}, status=400)
    if fecha_fin:
        try:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            asistencias_queryset = asistencias_queryset.filter(fecha__lte=fecha_fin)
        except ValueError:
            return JsonResponse({'error': 'Formato de fecha de fin inválido. Use YYYY-MM-DD.'}, status=400)

    asistencias_lista = []
    for asistencia in asistencias_queryset:
        asistencias_lista.append({
            'empleado': asistencia.empleado.cedula,
            'fecha': asistencia.fecha.strftime('%d-%m-%Y'),
            'hora_inicio': str(asistencia.hora_entrada),
            'hora_fin': str(asistencia.hora_salida),
            'estado': asistencia.estado,
            'observaciones': asistencia.observaciones,
        })

    return JsonResponse(asistencias_lista, safe=False)

@csrf_exempt
@require_http_methods(["GET"])
def get_faltas_justificables(request):
    try:
        cedula = request.GET.get('cedula')
        if not cedula:
            return JsonResponse({'error': 'Cédula del empleado es requerida'}, status=400)

        empleado_obj = empleado.objects.get(cedula=cedula)
        
        # Filtrar faltas no justificadas (que pueden ser justificadas)
        faltas = asistencias.objects.filter(
            empleado=empleado_obj,
            estado="F"  # Cambié a "F" (Falta) en lugar de "J" (Justificada)
        ).values('id', 'fecha', 'hora_entrada', 'hora_salida', 'observaciones')

        # Convertir el QuerySet a lista y retornar
        return JsonResponse(list(faltas), safe=False)
        
    except empleado.DoesNotExist:
        return JsonResponse({'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        # Captura cualquier otro error inesperado
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def listar_justificaciones(request):
    try:
        cedula = request.GET.get('cedula', '').strip()
        fecha_inicio = request.GET.get('fecha_inicio', '').strip()
        fecha_fin = request.GET.get('fecha_fin', '').strip()

        justificaciones_qs = Justificacion.objects.select_related(
            'asistencias__empleado',
            'aprobado_por'
        ).all()

        if cedula:
            justificaciones_qs = justificaciones_qs.filter(asistencias__empleado__cedula__icontains=cedula)

        if fecha_inicio:
            try:
                fecha_inicio_date = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                justificaciones_qs = justificaciones_qs.filter(asistencias__fecha__gte=fecha_inicio_date)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_inicio inválido. Use YYYY-MM-DD.'}, status=400)

        if fecha_fin:
            try:
                fecha_fin_date = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                justificaciones_qs = justificaciones_qs.filter(asistencias__fecha__lte=fecha_fin_date)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha_fin inválido. Use YYYY-MM-DD.'}, status=400)

        justificaciones_list = []
        for justificacion in justificaciones_qs:
            aprobado_por_nombre = None
            if justificacion.aprobado_por:
                aprobado_por_nombre = f"{justificacion.aprobado_por.primer_nombre} {justificacion.aprobado_por.primer_apellido}".strip()

            justificaciones_list.append({
                'id': justificacion.id,
                'tipo': justificacion.get_tipo_display(),
                'descripcion': justificacion.descripcion,
                'fecha_creacion': justificacion.fecha_creacion.strftime('%d-%m-%Y %H:%M'),
                'aprobado_por': aprobado_por_nombre,
                'fecha_aprobacion': justificacion.fecha_aprobacion.strftime('%d-%m-%Y %H:%M') if justificacion.fecha_aprobacion else None,
                'documento_url': justificacion.documento.url if justificacion.documento else None,
                'empleado_cedula': justificacion.asistencias.empleado.cedula if justificacion.asistencias and justificacion.asistencias.empleado else None,
                'fecha_asistencia': justificacion.asistencias.fecha.strftime('%d-%m-%Y') if justificacion.asistencias and justificacion.asistencias.fecha else None,
            })

        return JsonResponse(justificaciones_list, safe=False)

    except Exception as e:
        logger.error(f"Error en listar_justificaciones: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error interno del servidor', 'detail': str(e)}, status=500)

from django.utils.crypto import get_random_string

@csrf_exempt
@require_http_methods(["GET"])
def obtener_detalle_prenomina(request, prenomina_id):


    try:
        # Obtener la prenomina con sus detalles relacionados
        prenomina_obj = prenomina.objects.prefetch_related(
            'detalles__codigo',  # Prefetch para optimizar las consultas
        ).get(id_prenomina=prenomina_id)

        # Calcular el total de la prenomina
        total = prenomina_obj.detalles.aggregate(total=Sum('total_monto'))['total'] or 0

        # Preparar los datos de la prenomina
        prenomina_data = {
            'id_prenomina': prenomina_obj.id_prenomina,
            'periodo': prenomina_obj.nomina.periodo,
            'tipo': prenomina_obj.nomina.tipo_nomina.tipo_nomina,
            'secuencia': prenomina_obj.nomina.secuencia.nombre_secuencia,
            'fecha_creacion': prenomina_obj.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'total': total,
        }

        # Preparar los datos de los detalles de la prenomina
        detalles_data = []
        for detalle in prenomina_obj.detalles.all():
            concepto = detalle.codigo  # Acceso al objeto concepto_pago relacionado

            # Contar personas con este concepto en la nómina asociada
            numero_personas = detalle_nomina.objects.filter(
                nomina=prenomina_obj.nomina,
                codigo=concepto
            ).values('cedula').distinct().count()

            detalles_data.append({
                'codigo': concepto.codigo,
                'descripcion': concepto.descripcion,
                'tipo_concepto': concepto.tipo_concepto,
                'nombre_nomina': concepto.nombre_nomina,
                'asignacion': str(detalle.total_monto) if concepto.tipo_concepto == 'ASIGNACION' else '0',
                'deduccion': str(detalle.total_monto) if concepto.tipo_concepto == 'DEDUCCION' else '0',
                'total_monto': str(detalle.total_monto),
                'numero_personas': numero_personas,  # Nuevo campo agregado
            })

        # Retornar la respuesta en formato JSON
        return JsonResponse({
            'success': True,
            'prenomina': prenomina_data,
            'detalles': detalles_data,
        })

    except prenomina.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'La prenomina no existe'
        }, status=404)
    except Exception as e:
        logger.error(f"Error en obtener_detalle_prenomina: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)

#   <-------CREAR CUENTA------->

# Formulario verificacion empleado
def verificar_empleado(request):
    if request.method == 'POST':
        cedula = request.POST.get('formulario_cedula', '').strip()

        try:
            empleado_instance = empleado.objects.get(cedula__iexact=cedula)
            tiene_usuario = usuario.objects.filter(empleado=empleado_instance).exists()

            empleado_data = {
                'primer_nombre': empleado_instance.primer_nombre,
                'segundo_nombre': empleado_instance.segundo_nombre or '',
                'primer_apellido': empleado_instance.primer_apellido,
                'segundo_apellido': empleado_instance.segundo_apellido or '',
                'tiene_usuario': tiene_usuario,
                'cedula': cedula
            }

            return JsonResponse({
                'status': 'success',
                'tiene_usuario': tiene_usuario,
                'empleado': empleado_data,
                'message': 'Empleado verificado correctamente'
            })

        except empleado.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'No eres empleado registrado'
            }, status=404)

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': 'Error interno del servidor'
            }, status=500)

    return JsonResponse({
        'status': 'error',
        'message': 'Método no permitido'
    }, status=405)

# Solo si estás usando CSRF en AJAX, asegúrate de que esto sea seguro
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.hashers import make_password, check_password


#formnulario crear Cuenta
@require_POST
def crear_cuenta_empleado(request):
    try:
        # Verificar contraseñas (usa los mismos nombres que en el formulario)
        if request.POST.get('contraseña') != request.POST.get('confirmar_contraseña'):
            return JsonResponse({'error': 'Las contraseñas no coinciden'}, status=400)
        
        #Verificar que la contraseña sea mayour a 8 digitos
        if len(request.POST.get('contraseña', '')) < 8:
            return JsonResponse({'error': 'La contraseña debe tener al menos 8 caracteres'}, status=400)
            
        # Obtener datos
        email = request.POST.get('email')
        contraseña = request.POST.get('contraseña')
        cedula = request.POST.get('cedula')
        print(cedula)
        
        # Verificar email único
        if usuario.objects.filter(email=email).exists():
            return JsonResponse({'error': 'El email ya está registrado'}, status=400)
        contraseña_hash = make_password(contraseña)
        
        token_registro = get_random_string(50)
        
        # Almacenar usuario
        request.session['datos_registro'] = {
            'token': token_registro,
            'email': email,
            'contraseña': contraseña_hash,
            'cedula': cedula,
            'ultimo_login': str(timezone.now())
        }
        request.session.modified = True  # ¡Importante!
        request.session.set_expiry(3600)  # 1 hora de validez
        
        response = JsonResponse({
            'status': 'success',
            'message': 'Datos validados correctamente',
            'token': token_registro,
        })
        response.set_cookie(
            'token_registro', 
            token_registro, 
            max_age=3600, 
            httponly=False, 
            samesite='Lax',
            path='/',
        )
        
        return response        
        
    except empleado.DoesNotExist:
        return JsonResponse({'error': 'No existe un empleado con esta cédula'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error del servidor: {str(e)}'}, status=500)
    
#Completar registros para almacenar en la base de datos
def completar_registro(request):

    try:
        # 1. Verificación de token
        token_frontend = request.POST.get('token')
        datos_sesion = request.session.get('datos_registro', {})
        print("Token desde el frontend:", token_frontend)
        print("Datos de sesión:", datos_sesion)
        
        if not datos_sesion or token_frontend != datos_sesion.get('token'):
            return JsonResponse({'error': 'Sesión inválida o expirada'}, status=400)
        empleado_instance = empleado.objects.get(cedula=datos_sesion['cedula'])

        # 2. Validar preguntas/respuestas primero
        preguntas_respuestas = [
            (request.POST.get('pregunta1'), request.POST.get('respuesta1')),
            (request.POST.get('pregunta2'), request.POST.get('respuesta2')),
            (request.POST.get('pregunta3'), request.POST.get('respuesta3'))
        ]
        
        if any(not pr[0] or not pr[1] for pr in preguntas_respuestas):
            return JsonResponse({'error': 'Todas las preguntas deben tener respuesta'}, status=400)

        # 3. Crear usuario (SIN HASH)
        nuevo_usuario = usuario.objects.create(
            email=datos_sesion['email'],
            contraseña_hash=datos_sesion['contraseña'],  # ← Contraseña en texto plano
            empleado=empleado_instance,  # Usamos _id para asignación directa
            ultimo_login=timezone.now(),
        )
        rol_principal = rol.objects.get(codigo_rol= 1)
        roles_usuario = usuario_rol.objects.create(
            rol_id = rol_principal.codigo_rol,
            usuario_id = nuevo_usuario.id,
            fecha_asignacion = timezone.now()
        )

        # 4. Preguntas de seguridad (SIN HASH)
        for pregunta_id, respuesta in preguntas_respuestas:
            usuario_pregunta.objects.create(
                usuario_id=nuevo_usuario.id,  # Asignación por ID explícito
                pregunta_id=pregunta_id,
                respuesta_hash=make_password(respuesta.lower().strip()),  # ← Respuesta en texto plano
            )


        # 6. Limpiar sesión
        del request.session['datos_registro']

        return JsonResponse({
            'status': 'success',
            'message': 'Usuario registrado (modo pruebas sin hashing)!'
        })

    except Exception as e:
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)
    
#   <-------CREAR CUENTA TERMINADO------->

#          <-------LOGIN------->

@require_POST
def login_empleado(request):
    try:
        email = request.POST.get('email')
        contraseña = request.POST.get('password')

        # Validaciones de campos
        if not email:
            return JsonResponse({'status': 'error', 'error': 'El correo electrónico es obligatorio.'}, status=400)
        
        if not contraseña:
            return JsonResponse({'status': 'error', 'error': 'La contraseña es obligatoria.'}, status=400)

        # Buscar el usuario por email
        try:
            usuario_instance = usuario.objects.get(email=email)  # Corregí el typo de 'sle' a 'get'
        except usuario.DoesNotExist:
            return JsonResponse({'status': 'error', 'error': 'Correo electrónico no registrado.'}, status=401)

        # Verificar contraseña
        if not check_password(contraseña, usuario_instance.contraseña_hash):
            return JsonResponse({'status': 'error', 'error': 'Contraseña incorrecta.'}, status=401)

        # Obtener el empleado asociado
        empleado_instance = usuario_instance.empleado

        # Guardar datos de sesión
        request.session['usuario_id'] = usuario_instance.id
        request.session['email'] = usuario_instance.email
        request.session['empleado_id'] = usuario_instance.empleado_id
        request.session.set_expiry(3600)  # 1 hora

        usuario_instance.ultimo_login = timezone.now()
        usuario_instance.save()

        # Obtener el nombre del rol como array para mantener compatibilidad con el frontend
        usuario_roles = usuario_rol.objects.filter(usuario=usuario_instance).select_related('rol')
        if usuario_roles.exists():
            rol_nombre = [ur.rol.nombre_rol for ur in usuario_roles]
        else:
            rol_nombre = ['Sin rol asignado']

        # Información del usuario para enviar al frontend
        usuario_info = {
            'nombre': empleado_instance.primer_nombre,
            'apellido': empleado_instance.primer_apellido,
            'cedula': empleado_instance.cedula,
            'cargo': str(empleado_instance.cargo) if empleado_instance.cargo else 'N/A',
            'roles': rol_nombre,  # Ahora es un array
            'ultimo_login': usuario_instance.ultimo_login.strftime('%d/%m/%Y %H:%M') if usuario_instance.ultimo_login else 'Nunca'
        }
        
        return JsonResponse({
            'status': 'success',
            'message': 'Inicio de sesión exitoso.',
            'redirect_url': '/menu/',
            'usuario_info': usuario_info
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'error': str(e)}, status=500)
#          <-------LOGIN_TERMINADO------->


from django.contrib.auth import logout
from django.views.decorators.http import require_POST
from django.http import JsonResponse

@require_POST
def logout_empleado(request):
    try:
        # Limpiar la sesión
        logout(request)  # Esto elimina la sesión de autenticación
        request.session.flush()  # Limpia todos los datos de la sesión
        
        return JsonResponse({
            'status': 'success',
            'message': 'Sesión cerrada correctamente',
            'redirect_url': '/login/'  # Redirige a la página de login
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': str(e)
        }, status=500)

#          <-------RECUPERAR_CONTRASEÑA------->

"""class CustomLoginView(APIView):
    permission_classes = [AllowAny]
    http_method_names = ['post', 'get']
    
    def get(self, request):
        return render(request, 'login.html')

    def post(self, request):
        email = request.data.get("email")
        password = request.data.get("password")
        print("Correo recibido:", email)
        print("Datos recibidos en backend:", email, password)  # Verificación

        try:
            user = usuario.objects.get(email=email)
        except usuario.DoesNotExist:
            return Response({"error": "Email no registrado"}, status=HTTP_400_BAD_REQUEST)

        if user.check_password(password):
            refresh = RefreshToken.for_user(user)
            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
            }, status=HTTP_200_OK)
        else:
            return Response({"error": "Contraseña incorrecta"}, status=HTTP_400_BAD_REQUEST)"""

#   <----------codigo importar documento -------------------->


@csrf_exempt
@require_http_methods(["POST"])
def crear_justificacion(request):
    if request.method == 'POST':
        try:
            # Obtener datos del formulario (usando los nombres correctos)
            falta_id = request.POST.get('falta')
            tipo = request.POST.get('tipo')  # Nombre correcto del select
            descripcion = request.POST.get('motivo')  # Nombre correcto del textarea
            documento = request.FILES.get('documento')

            # Validación modificada (el documento es opcional)
            if not all([falta_id, tipo]):
                return JsonResponse({
                    'success': False, 
                    'message': 'Falta ID y Tipo son campos requeridos.'
                }, status=400)

            # Obtener la asistencia
            try:
                asistencia_obj = asistencias.objects.get(pk=falta_id)
            except asistencias.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Asistencia no encontrada.'
                }, status=404)

            # Actualizar la asistencia
            asistencia_obj.estado = "J"
            if descripcion:
                asistencia_obj.observaciones = descripcion
            asistencia_obj.save()

            # Crear la justificación
            justificacion = Justificacion.objects.create(
                asistencias=asistencia_obj,
                tipo=tipo,
                descripcion=descripcion or "",  # Asegurar que no sea None
                documento=documento
            )

            return JsonResponse({
                'success': True,
                'message': 'Justificación registrada correctamente',
                'justificacion_id': justificacion.id
            }, status=201)

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error del servidor: {str(e)}'
            }, status=500)
        
from django.shortcuts import get_object_or_404
from django.db.models import Prefetch

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import recibo_pago

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Sum
from .models import recibo_pago
import logging

logger = logging.getLogger(__name__)

@require_GET
def obtener_datos_recibo(request, recibo_id):
    try:
        # Optimización de consultas con select_related y prefetch_related
        recibo = (recibo_pago.objects
                .select_related(
                    'cedula',
                    'nomina',
                    'cedula__cargo',
                    'cedula__cargo__familia',
                    'cedula__cargo__nivel'
                )
                .prefetch_related(
                    'detalles__detalle_nomina__codigo',
                    'cedula__cuentas_bancarias__banco'
                )
                .get(pk=recibo_id))

        # Obtener cuenta bancaria activa
        cuenta_activa = recibo.cedula.cuentas_bancarias.filter(activa=True).first()

        # Procesar conceptos
        conceptos = []
        for detalle in recibo.detalles.all():
            concepto = detalle.detalle_nomina
            conceptos.append({
                'codigo': concepto.codigo.codigo,
                'descripcion': concepto.codigo.descripcion,
                'asignacion': float(concepto.monto) if concepto.codigo.tipo_concepto == 'ASIGNACION' else None,
                'deduccion': float(concepto.monto) if concepto.codigo.tipo_concepto == 'DEDUCCION' else None
            })

        # Cálculo de totales
        detalles = recibo.detalles.all()
        total_asignaciones = sum(d.detalle_nomina.monto for d in detalles if d.detalle_nomina.codigo.tipo_concepto == 'ASIGNACION')
        total_deducciones = sum(d.detalle_nomina.monto for d in detalles if d.detalle_nomina.codigo.tipo_concepto == 'DEDUCCION')

        datos = {
            'encabezado': {
                'fecha_ingreso': recibo.cedula.fecha_ingreso.strftime("%d-%b-%y") if recibo.cedula.fecha_ingreso else "N/A",
                'nombre_completo': recibo.cedula.get_nombre_completo(),
                'cedula': recibo.cedula.cedula,
                'numero_cuenta': cuenta_activa.numero_cuenta if cuenta_activa else "No registrado",
                'banco': cuenta_activa.banco.nombre if cuenta_activa else "No especificado",
                'cargo': recibo.cedula.cargo.nombre_completo if recibo.cedula.cargo else "Sin cargo",
                'periodo': recibo.nomina.periodo if recibo.nomina else "Periodo no definido",
                'sueldo_base': next((c['asignacion'] for c in conceptos if c['codigo'] == '1001'), 0)  # Asumiendo que 1001 es sueldo base
            },
            'conceptos': conceptos,
            'totales': {
                'total_asignaciones': total_asignaciones,
                'total_deducciones': total_deducciones,
                'total_nomina': total_asignaciones - total_deducciones
            }
        }
        
        return JsonResponse(datos)

    except recibo_pago.DoesNotExist:
        return JsonResponse({'error': 'Recibo no encontrado'}, status=404)
    except Exception as e:
        logger.error(f"Error al obtener recibo {recibo_id}: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Error interno del servidor'}, status=500)
    
@login_required
def listado_recibos(request):
    try:
        # Obtener el usuario personalizado
        try:
            usuario_personalizado = usuario.objects.get(email=request.user.email)
        except usuario.DoesNotExist:
            return JsonResponse({'error': 'Usuario no encontrado'}, status=404)
        
        # Verificar que tenga empleado asociado
        if not usuario_personalizado.empleado:
            return JsonResponse({'error': 'Usuario no tiene empleado asociado'}, status=400)
        
        recibos = recibo_pago.objects.filter(
            cedula=usuario_personalizado.empleado
        ).select_related(
            'cedula',
            'cedula__cargo',
            'nomina'
        ).order_by('-fecha_generacion')
        
        paginator = Paginator(recibos, 10)
        page_number = request.GET.get('page')
        
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
            
        return render(request, 'menu_principal/subs_menus/recibo_pago.html', {
            'page_obj': page_obj,
            'titulo_pagina': 'Mis Recibos de Pago'
        })
        
    except Exception as e:
        print(f"Error en listado_recibos: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
    


logger = logging.getLogger(__name__)

def generar_prenomina_para_nomina(nomina_instance):
    try:
        if prenomina.objects.filter(nomina=nomina_instance).exists():
            logger.info(f"Prenomina ya existe para la nómina {nomina_instance.id_nomina}")
            return None

        nueva = prenomina.objects.create(nomina=nomina_instance, fecha_creacion=now())

        detalles = detalle_nomina.objects.filter(nomina=nomina_instance)
        if not detalles.exists():
            logger.warning(f"No hay detalles para la nómina {nomina_instance.id_nomina}")
            return None

        resumen = detalles.values('codigo').annotate(total=Sum('monto'))
        for item in resumen:
            detalle_prenomina.objects.create(
                prenomina=nueva, codigo_id=item['codigo'], total_monto=item['total']
            )

        logger.info(f"Prenomina generada para la nómina {nomina_instance.id_nomina}")
        return nueva

    except Exception as e:
        logger.error(f"Error generando prenomina {nomina_instance.id_nomina}: {e}", exc_info=True)
        raise



#   /////////////// Importar nominas al panel ///////////////

@require_http_methods(["GET"])
def listar_nominas(request):
    """API para listar nóminas con filtros mejorados"""
    try:
        # 1. Obtener y validar parámetros de filtrado
        tipo = request.GET.get('tipo', '').strip()
        mes = request.GET.get('mes', '').strip()
        anio = request.GET.get('anio', '').strip()
        estado = request.GET.get('estado', '').strip()  # Nuevo filtro por estado
        orden = request.GET.get('orden', '-fecha_carga')
        
        # 2. Construir consulta base con select_related para optimización
        queryset = nomina.objects.select_related('tipo_nomina', 'secuencia').all()
        
        # 3. Aplicar filtros con condiciones más precisas
        filters = Q()
        
        if tipo:
            filters &= Q(tipo_nomina__tipo_nomina__icontains=tipo)
        
        if mes:
            # Asume que periodo tiene formato "MM-YYYY" o similar
            if len(mes) == 1:
                mes = f'0{mes}'  # Normalizar a dos dígitos
            filters &= Q(periodo__contains=f'-{mes}-') | Q(periodo__startswith=f'{mes}-')
        
        if anio:
            # Busca año al inicio (2023-), en medio (-2023-) o al final (-2023)
            filters &= Q(periodo__contains=anio)
            
        if estado:  # Nuevo filtro por estado
            filters &= Q(estado__iexact=estado)
        
        queryset = queryset.filter(filters)
        
        # 4. Validar y aplicar ordenamiento
        campos_orden_validos = [
            '-fecha_carga', 'fecha_carga',
            'tipo_nomina__tipo_nomina', '-tipo_nomina__tipo_nomina',
            '-fecha_cierre', 'fecha_ciere',
            'estado', '-estado'  # Nuevos campos de ordenamiento
        ]
        
        if orden not in campos_orden_validos:
            orden = '-fecha_carga'
            
        queryset = queryset.order_by(orden)
        
        # 5. Paginación con manejo de errores
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 25)  # 25 items por página
        
        try:
            nominas_paginadas = paginator.page(page_number)
        except:
            nominas_paginadas = paginator.page(1)  # Fallback a primera página
        
        # 6. Preparar datos para respuesta con manejo de valores nulos
        nominas_data = []
        for nom in nominas_paginadas:
            nominas_data.append({
                'id_nomina': nom.id_nomina,
                'tipo_nomina': nom.tipo_nomina.tipo_nomina if nom.tipo_nomina else '',
                'periodo': nom.periodo,
                'secuencia': nom.secuencia.nombre_secuencia if nom.secuencia else '',
                'fecha_cierre': nom.fecha_cierre.strftime('%d/%m/%Y') if nom.fecha_cierre else '',
                'fecha_carga': nom.fecha_carga.strftime('%d/%m/%Y %H:%M') if nom.fecha_carga else '',
                'estado': nom.estado,  # Nuevo campo estado
                'estado_display': nom.get_estado_display(),  # Versión legible del estado
                'total_empleados': detalle_nomina.objects.filter(nomina=nom).values('cedula').distinct().count(),
                'total_conceptos': detalle_nomina.objects.filter(nomina=nom).count(),
                'fecha_aprobacion': nom.fecha_aprobacion.strftime('%d/%m/%Y %H:%M') if nom.fecha_aprobacion else None
            })
        
        # 7. Retornar respuesta estructurada
        return JsonResponse({
            'success': True,
            'nominas': nominas_data,
            'total': paginator.count,
            'paginas': paginator.num_pages,
            'actual': nominas_paginadas.number,
            'params': {  # Para debugging
                'tipo': tipo,
                'mes': mes,
                'anio': anio,
                'estado': estado,  # Nuevo parámetro
                'orden': orden
            }
        })
        
    except Exception as e:
        logger.error(f"Error en listar_nominas: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)

from django.contrib.auth.decorators import login_required
from django.db.models import Sum

logger = logging.getLogger(__name__)

@csrf_exempt
@transaction.atomic
@require_http_methods(["GET"])
def listar_prenominas(request):
    """API para listar prenominas con filtros mejorados"""
    try:
        # 1. Obtener y validar parámetros de filtrado
        tipo = request.GET.get('tipo', '').strip()
        mes = request.GET.get('mes', '').strip()
        anio = request.GET.get('anio', '').strip()
        orden = request.GET.get('orden', '-fecha_creacion')
        
        # 2. Construir consulta base con select_related para optimización
        queryset = prenomina.objects.select_related('nomina', 'nomina__tipo_nomina', 'nomina__secuencia').all()
        
        # 3. Aplicar filtros con condiciones más precisas
        filters = Q()
        
        if tipo:
            filters &= Q(nomina__tipo_nomina__tipo_nomina__icontains=tipo)
        
        if mes:
            # Asume que periodo tiene formato "MM-YYYY" o similar
            if len(mes) == 1:
                mes = f'0{mes}'  # Normalizar a dos dígitos
            filters &= Q(nomina__periodo__contains=f'-{mes}-') | Q(nomina__periodo__startswith=f'{mes}-')
        
        if anio:
            # Busca año al inicio (2023-), en medio (-2023-) o al final (-2023)
            filters &= Q(nomina__periodo__contains=anio)
        
        queryset = queryset.filter(filters)
        
        # 4. Validar y aplicar ordenamiento
        campos_orden_validos = [
            '-fecha_creacion', 'fecha_creacion',
            'nomina__tipo_nomina__tipo_nomina', '-nomina__tipo_nomina__tipo_nomina',
            '-nomina__fecha_cierre', 'nomina__fecha_cierre'
        ]
        
        if orden not in campos_orden_validos:
            orden = '-fecha_creacion'
            
        queryset = queryset.order_by(orden)
        
        # 5. Paginación con manejo de errores
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 25)  # 25 items por página
        
        try:
            prenominas_paginadas = paginator.page(page_number)
        except:
            prenominas_paginadas = paginator.page(1)  # Fallback a primera página
        
        # 6. Preparar datos para respuesta con manejo de valores nulos
        prenominas_data = []
        for prenomina_obj in prenominas_paginadas:
            # Calcular el total de la prenomina
            total = prenomina_obj.detalles.aggregate(total=Sum('total_monto'))['total'] or 0
            
            prenominas_data.append({
                'id_prenomina': prenomina_obj.id_prenomina,
                'tipo_nomina': prenomina_obj.nomina.tipo_nomina.tipo_nomina if prenomina_obj.nomina.tipo_nomina else '',
                'periodo': prenomina_obj.nomina.periodo if prenomina_obj.nomina else '',
                'secuencia': prenomina_obj.nomina.secuencia.nombre_secuencia if prenomina_obj.nomina.secuencia else '',
                'fecha_cierre': prenomina_obj.nomina.fecha_cierre.strftime('%d/%m/%Y') if prenomina_obj.nomina.fecha_cierre else '',
                'fecha_creacion': prenomina_obj.fecha_creacion.strftime('%d/%m/%Y %H:%M') if prenomina_obj.fecha_creacion else '',
                'total': total
            })
        
        # 7. Retornar respuesta estructurada
        return JsonResponse({
            'success': True,
            'prenominas': prenominas_data,
            'total': paginator.count,
            'paginas': paginator.num_pages,
            'actual': prenominas_paginadas.number,
            'params': {  # Para debugging
                'tipo': tipo,
                'mes': mes,
                'anio': anio,
                'orden': orden
            }
        })
        
    except Exception as e:
        logger.error(f"Error en listar_prenominas: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)

@csrf_exempt
@transaction.atomic
def eliminar_nomina(request, pk):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            if data.get('action') == 'delete':
                # Intenta obtener y eliminar la nómina
                try:
                    nomina_obj = nomina.objects.get(id_nomina=pk)
                    
                    # Eliminación en cascada
                    detalle_recibo.objects.filter(detalle_nomina__nomina=nomina_obj).delete()
                    recibo_pago.objects.filter(nomina=nomina_obj).delete()
                    
                    try:
                        prenomina_obj = prenomina.objects.get(nomina=nomina_obj)
                        detalle_prenomina.objects.filter(prenomina=prenomina_obj).delete()
                        prenomina_obj.delete()
                    except prenomina.DoesNotExist:
                        pass
                        
                    detalle_nomina.objects.filter(nomina=nomina_obj).delete()
                    nomina_obj.delete()
                    
                    return JsonResponse({
                        'status': 'success',
                        'message': 'Nómina eliminada correctamente'
                    }, status=200)
                
                except nomina.DoesNotExist:
                    # Si la nómina no existe, considerarlo como éxito (quizás ya fue eliminada)
                    return JsonResponse({
                        'status': 'success',
                        'message': 'La nómina ya no existe (posiblemente eliminada)'
                    }, status=200)
                    
        except Exception as e:
            return JsonResponse({
                'error': str(e),
                'details': 'Error al procesar la solicitud'
            }, status=500)
    
    return JsonResponse({
        'error': 'Método no permitido'
    }, status=405)

#   /////////////// Importar empleados al panel ///////////////

@require_http_methods(["GET"])
def listar_empleados(request):
    """API para listar empleados con todos los campos requeridos"""
    try:
        # 1. Obtener y validar parámetros de filtrado
        cedula = request.GET.get('cedula', '').strip()
        nombre = request.GET.get('nombre', '').strip()
        apellido = request.GET.get('apellido', '').strip()
        cargo = request.GET.get('cargo', '').strip()
        tipo_trabajador = request.GET.get('tipo_trabajador', '').strip()
        status = request.GET.get('status', '').strip()
        orden = request.GET.get('orden', '-fecha_ingreso')
        
        # 2. Construir consulta base optimizada
        queryset = empleado.objects.select_related(
            'cargo',
            'cargo__familia',
            'cargo__nivel',
            'tipo_trabajador'
        ).prefetch_related(
            'cuentas_bancarias',
            'cuentas_bancarias__banco'
        ).all()
        
        # 3. Aplicar filtros
        filters = Q()
        
        if cedula:
            filters &= Q(cedula__icontains=cedula)
        
        if nombre:
            filters &=  (Q(primer_nombre__icontains=nombre) | 
                        Q(segundo_nombre__icontains=nombre))
        
        if apellido:
            filters &=  (Q(primer_apellido__icontains=apellido) | 
                        Q(segundo_apellido__icontains=apellido))
        
        if cargo:
            # Buscar en familia.nombre o nivel.nivel
            filters &=  (Q(cargo__familia__nombre__icontains=cargo) |
                        Q(cargo__nivel__nivel__icontains=cargo))
        
        if tipo_trabajador:
            filters &= Q(tipo_trabajador__descripcion__icontains=tipo_trabajador)
        
        if status:
            if status.lower() == 'true':
                filters &= Q(status=True)
            elif status.lower() == 'false':
                filters &= Q(status=False)
        
        queryset = queryset.filter(filters)
        
        # 4. Validar y aplicar ordenamiento
        campos_orden_validos = [
            '-fecha_ingreso', 'fecha_ingreso',
            'primer_apellido', '-primer_apellido',
            'cargo__familia__nombre', '-cargo__familia__nombre',
            'cargo__nivel__orden_jerarquico', '-cargo__nivel__orden_jerarquico',
            'cedula', '-cedula',
            'fecha_nacimiento', '-fecha_nacimiento'
        ]
        
        if orden not in campos_orden_validos:
            orden = '-fecha_ingreso'
            
        queryset = queryset.order_by(orden)
        
        # 5. Paginación
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 25)  # 25 items por página
        
        try:
            empleados_paginados = paginator.page(page_number)
        except:
            empleados_paginados = paginator.page(1)
        
        # 6. Preparar datos con TODOS los campos requeridos
        empleados_data = []
        for emp in empleados_paginados:
            # Formatear cuentas bancarias
            cuentas_info = []
            for cuenta in emp.cuentas_bancarias.all():
                if cuenta.activa:
                    cuentas_info.append(
                        f"{cuenta.banco.nombre} ({cuenta.get_tipo_display()}): {cuenta.numero_cuenta}"
                    )
            
            # Obtener nombre del cargo (familia + nivel)
            nombre_cargo = ""
            if emp.cargo:
                nombre_cargo = f"{emp.cargo.familia.nombre} - Nivel {emp.cargo.nivel.nivel}"
            
            empleados_data.append({
                # Identificación
                'tipo_id': emp.get_tipo_identificacion_display(),
                'cedula': emp.cedula,
                
                # Nombres
                'nombre_completo': f"{emp.primer_nombre} {emp.segundo_nombre or ''} {emp.primer_apellido} {emp.segundo_apellido or ''}".strip(),
                
                # Datos personales
                'fecha_nacimiento': emp.fecha_nacimiento.strftime('%d/%m/%Y') if emp.fecha_nacimiento else '',
                'lugar_nacimiento': emp.lugar_nacimiento or '',
                'genero': emp.get_genero_display() if emp.genero else '',
                'estado_civil': emp.get_estado_civil_display() if emp.estado_civil else '',
                
                # Datos laborales
                'fecha_ingreso': emp.fecha_ingreso.strftime('%d/%m/%Y') if emp.fecha_ingreso else '',
                'cargo': nombre_cargo,  # Usamos la combinación familia + nivel
                'tipo_trabajador': emp.tipo_trabajador.descripcion if emp.tipo_trabajador else '',
                'grado_instruccion': emp.grado_instruccion or '',
                
                # Contacto
                'telefono_principal': emp.telefono_principal or '',
                'telefono_secundario': emp.telefono_secundario or '',
                'email': emp.email or '',
                'direccion': emp.direccion or '',
                
                # Datos familiares
                'hijos': emp.hijos,
                'conyuge': 'Sí' if emp.conyuge else 'No',
                
                # Información adicional
                'rif': emp.rif or '',
                
                # Cuentas bancarias (formateadas como texto)
                'cuentas_bancarias': "\n".join(cuentas_info) if cuentas_info else 'Sin cuentas activas',
                
                # Estado
                'status': 'Activo' if emp.status else 'Inactivo',
                
                # ID para acciones
                'id': emp.cedula
            })
        
        # 7. Retornar respuesta
        return JsonResponse({
            'success': True,
            'empleados': empleados_data,
            'total': paginator.count,
            'paginas': paginator.num_pages,
            'actual': empleados_paginados.number,
            'params': {
                'cedula': cedula,
                'nombre': nombre,
                'apellido': apellido,
                'cargo': cargo,
                'tipo_trabajador': tipo_trabajador,
                'status': status,
                'orden': orden
            }
        })
        
    except Exception as e:
        logger.error(f"Error en listar_empleados: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)
    

from django.core.exceptions import ValidationError
from django.db import IntegrityError

def empleado_view(request):
    """
    Vista para renderizar el formulario de creación de empleados
    """
    tipos_trabajador = tipo_trabajador.objects.all().order_by('descripcion')
    cargos_completos = cargo.objects.select_related('familia', 'nivel').filter(activo=True).order_by('familia__nombre', 'nivel__orden_jerarquico')
    bancos = banco.objects.all().order_by('nombre')

    context = {
        'tipos_trabajador': tipos_trabajador,
        'cargos_completos': cargos_completos,
        'bancos': bancos
    }
    
    return render(request, 'empleados/crear_empleado.html', context)

logger = logging.getLogger(__name__)
@csrf_exempt
@transaction.atomic
def api_empleadoss(request):
    """
    API para crear nuevos empleados con sus hijos y cuentas bancarias
    """
    if request.method == 'POST':
        try:
            # 1. Parsear y validar JSON de entrada
            try:
                data = json.loads(request.body)
                logger.info(f"Datos recibidos: {data}")
            except json.JSONDecodeError as e:
                logger.error("Error decodificando JSON: %s", str(e))
                return JsonResponse({
                    'success': False,
                    'error': 'Formato JSON inválido',
                    'detail': str(e)
                }, status=400)

            # 2. Validar campos requeridos del empleado
            required_fields = [
                'tipo_identificacion', 'cedula', 'primer_nombre',
                'primer_apellido', 'fecha_ingreso', 'tipo_trabajador',
                'cargo_id'
            ]
            
            if missing_fields := [field for field in required_fields if field not in data]:
                error_msg = f'Campos requeridos faltantes: {", ".join(missing_fields)}'
                logger.error(error_msg)
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)

            # 3. Convertir y validar tipos de datos
            try:
                # Conversiones básicas
                data['cedula'] = int(data['cedula'])
                data['tipo_trabajador'] = int(data['tipo_trabajador'])
                data['cargo_id'] = int(data['cargo_id'])
                data['hijos'] = int(data.get('hijos', 0))
                
                # Nivel salarial (opcional)
                if 'nivel_salarial' in data and data['nivel_salarial']:
                    data['nivel_salarial'] = int(data['nivel_salarial'])
                
                # Procesamiento de fechas
                fecha_nacimiento = (datetime.strptime(data['fecha_nacimiento'], '%Y-%m-%d').date() 
                                if data.get('fecha_nacimiento') else None)
                fecha_ingreso = datetime.strptime(data['fecha_ingreso'], '%Y-%m-%d').date()
                
                # Validar hijos_data si existe
                hijos_data = data.get('hijos_data', [])
                if not isinstance(hijos_data, list):
                    raise ValueError("hijos_data debe ser una lista")
                    
            except ValueError as e:
                logger.error("Error en conversión de datos: %s", str(e))
                return JsonResponse({
                    'success': False,
                    'error': 'Error en tipos de datos',
                    'detail': str(e)
                }, status=400)

            # 4. Verificar existencia de relaciones
            if not tipo_trabajador.objects.filter(codigo_trabajador=data['tipo_trabajador']).exists():
                error_msg = f"Tipo de trabajador {data['tipo_trabajador']} no existe"
                logger.error(error_msg)
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)

            # 5. Validar nivel salarial si fue proporcionado
            if data.get('nivel_salarial') and not nivel_salarial.objects.filter(id=data['nivel_salarial']).exists():
                error_msg = f"Nivel salarial {data['nivel_salarial']} no existe"
                logger.error(error_msg)
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)

            # 6. Obtener y validar el cargo
            try:
                cargo_obj = cargo.objects.select_related('familia__tipo_trabajador').get(id=data['cargo_id'])
                
                if cargo_obj.familia.tipo_trabajador.codigo_trabajador != data['tipo_trabajador']:
                    error_msg = "El cargo no corresponde al tipo de trabajador seleccionado"
                    logger.error(error_msg)
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    }, status=400)
                
                if not cargo_obj.activo:
                    cargo_obj.activo = True
                    cargo_obj.save()
                    
            except cargo.DoesNotExist:
                error_msg = f"Cargo con ID {data['cargo_id']} no existe"
                logger.error(error_msg)
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)

            # 7. Verificar unicidad de datos
            if empleado.objects.filter(cedula=data['cedula']).exists():
                error_msg = f"Ya existe un empleado con cédula {data['cedula']}"
                logger.error(error_msg)
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)

            if data.get('email') and empleado.objects.filter(email=data['email']).exists():
                error_msg = f"El email {data['email']} ya está registrado"
                logger.error(error_msg)
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                }, status=400)

            # 8. Preparar datos para crear el empleado
            empleado_data = {
                'tipo_identificacion': data['tipo_identificacion'],
                'cedula': data['cedula'],
                'rif': data.get('rif'),
                'primer_nombre': data['primer_nombre'],
                'segundo_nombre': data.get('segundo_nombre'),
                'primer_apellido': data['primer_apellido'],
                'segundo_apellido': data.get('segundo_apellido'),
                'fecha_nacimiento': fecha_nacimiento,
                'lugar_nacimiento': data.get('lugar_nacimiento'),
                'genero': data.get('genero'),
                'estado_civil': data.get('estado_civil'),
                'grado_instruccion': data.get('grado_instruccion'),
                'fecha_ingreso': fecha_ingreso,
                'tipo_trabajador_id': data['tipo_trabajador'],
                'cargo': cargo_obj,
                'nivel_salarial_id': data.get('nivel_salarial'),
                'telefono_principal': data.get('telefono_principal'),
                'telefono_secundario': data.get('telefono_secundario'),
                'email': data.get('email'),
                'direccion': data.get('direccion'),
                'hijos': data['hijos'],
                'conyuge': bool(data.get('conyuge', False)),
                'status': True
            }

            # 9. Crear el empleado
            try:
                nuevo_empleado = empleado.objects.create(**empleado_data)
                logger.info(f"Empleado creado: {nuevo_empleado.cedula}")
            except Exception as e:
                logger.error("Error creando empleado: %s", str(e))
                return JsonResponse({
                    'success': False,
                    'error': 'Error al crear empleado',
                    'detail': str(e)
                }, status=400)

            # 10. Procesar hijos del empleado
            hijos_creados = 0
            for hijo_data in hijos_data:
                try:
                    # Validar campos requeridos para hijo
                    required_hijo_fields = ['nombre_completo', 'fecha_nacimiento', 'genero']
                    if missing_hijo := [f for f in required_hijo_fields if f not in hijo_data]:
                        raise ValueError(f'Campos faltantes en hijo: {", ".join(missing_hijo)}')
                    
                    # Limpiar y formatear datos
                    hijo_data['fecha_nacimiento'] = datetime.strptime(
                        hijo_data['fecha_nacimiento'], '%Y-%m-%d'
                    ).date()
                    
                    if 'cedula' in hijo_data and hijo_data['cedula'] == '':
                        hijo_data['cedula'] = None
                    
                    # Crear registro del hijo
                    hijo.objects.create(
                        empleado=nuevo_empleado,
                        nombre_completo=hijo_data['nombre_completo'],
                        cedula=hijo_data.get('cedula'),
                        fecha_nacimiento=hijo_data['fecha_nacimiento'],
                        lugar_nacimiento=hijo_data.get('lugar_nacimiento'),
                        genero=hijo_data['genero'],
                        estudia=hijo_data.get('estudia', 'S'),
                        nivel_educativo=hijo_data.get('nivel_educativo'),
                        discapacidad=hijo_data.get('discapacidad', False)
                    )
                    hijos_creados += 1
                    logger.info(f"Hijo creado para empleado {nuevo_empleado.cedula}")
                    
                except Exception as e:
                    logger.error("Error creando hijo: %s - Datos: %s", str(e), hijo_data)
                    transaction.set_rollback(True)
                    return JsonResponse({
                        'success': False,
                        'error': f'Error creando hijo: {str(e)}',
                        'hijo_data': hijo_data
                    }, status=400)

            # 11. Crear cuenta bancaria si se proporcionan datos
            if all(k in data for k in ['banco', 'tipo_cuenta', 'numero_cuenta']):
                try:
                    if not banco.objects.filter(codigo=data['banco']).exists():
                        error_msg = f"Banco con código {data['banco']} no existe"
                        logger.error(error_msg)
                        return JsonResponse({
                            'success': False,
                            'error': error_msg
                        }, status=400)

                    cuenta_bancaria.objects.create(
                        empleado=nuevo_empleado,
                        banco_id=data['banco'],
                        tipo=data['tipo_cuenta'],
                        numero_cuenta=data['numero_cuenta'],
                        activa=True
                    )
                    logger.info(f"Cuenta bancaria creada para empleado {nuevo_empleado.cedula}")
                except IntegrityError:
                    error_msg = f"El número de cuenta {data['numero_cuenta']} ya existe"
                    logger.error(error_msg)
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    }, status=400)
                except Exception as e:
                    logger.error("Error creando cuenta bancaria: %s", str(e))
                    return JsonResponse({
                        'success': False,
                        'error': 'Error al crear cuenta bancaria',
                        'detail': str(e)
                    }, status=400)

            # 12. Retornar respuesta exitosa
            return JsonResponse({
                'success': True,
                'message': 'Empleado creado correctamente',
                'empleado_id': nuevo_empleado.cedula,
                'nombre_completo': f"{nuevo_empleado.primer_nombre} {nuevo_empleado.primer_apellido}",
                'hijos_creados': hijos_creados,
                'total_hijos': len(hijos_data)
            })

        except ValidationError as e:
            logger.error("Error de validación: %s", str(e))
            return JsonResponse({
                'success': False,
                'error': 'Error de validación',
                'detail': str(e)
            }, status=400)
            
        except Exception as e:
            logger.error("Error interno del servidor: %s", str(e), exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Error interno del servidor',
                'detail': str(e)
            }, status=500)
    
    return JsonResponse({
        'error': 'Método no permitido'
    }, status=405)

def api_cargos_por_tipo(request):
    """
    API para obtener cargos filtrados por tipo de trabajador
    """
    tipo_trabajador_id = request.GET.get('tipo_trabajador')
    
    if not tipo_trabajador_id:
        return JsonResponse([], safe=False)
    
    try:
        cargos = cargo.objects.filter(
            familia__tipo_trabajador_id=tipo_trabajador_id,
            activo=True
        ).select_related('familia', 'nivel').order_by('familia__nombre', 'nivel__orden_jerarquico')
        
        cargos_data = [{
            'id': c.id,
            'familia_nombre': c.familia.nombre,
            'nivel_nombre': c.nivel.nombre,
            'codigo': f"{c.familia.nombre} - {c.nivel.nombre}"
        } for c in cargos]
        
        return JsonResponse(cargos_data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

    from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
import json

@require_http_methods(["DELETE"])
@csrf_exempt
def eliminar_empleado(request, pk):
    if request.method == 'DELETE':
        try:
            # Buscar el empleado a eliminar (usando minúsculas como está definido)
            empleado_obj = empleado.objects.get(pk=pk)
            
            # Eliminar primero las cuentas bancarias asociadas
            cuenta_bancaria.objects.filter(empleado=empleado_obj).delete()
            
            # Eliminar el empleado
            empleado_obj.delete()
            
            # Respuesta de éxito
            return JsonResponse({
                'message': 'Empleado y cuentas bancarias asociadas eliminadas correctamente',
                'status': 'success'
            }, status=200)
            
        except empleado.DoesNotExist:
            return JsonResponse({
                'error': 'El empleado no existe',
                'status': 'error'
            }, status=404)
            
        except Exception as e:
            return JsonResponse({
                'error': str(e),
                'status': 'error'
            }, status=500)
        

#   /////////////// Panel de roles ///////////////

@require_http_methods(["GET"])
def listar_usuarios(request):
    """API para listar usuarios con sus roles y filtros"""
    try:
        # 1. Obtener parámetros
        nombre = request.GET.get('nombre', '').strip()
        email = request.GET.get('email', '').strip()
        rol_filter = request.GET.get('rol', '').strip()
        estado = request.GET.get('estado', '').strip()
        orden = request.GET.get('orden', '-fecha_registro')
        page = request.GET.get('page', 1)
        
        # 2. Construir consulta base con prefetch para roles
        queryset = usuario.objects.select_related('empleado').prefetch_related(
            Prefetch('usuario_rol_set', queryset=usuario_rol.objects.select_related('rol'))
        ).all()
        
        # 3. Aplicar filtros
        if nombre:
            queryset = queryset.filter(
                Q(empleado__primer_nombre__icontains=nombre) |
                Q(empleado__primer_apellido__icontains=nombre)
            )
        if email:
            queryset = queryset.filter(email__icontains=email)
        if rol_filter:
            queryset = queryset.filter(usuario_rol__rol__nombre_rol__icontains=rol_filter).distinct()
        if estado:
            queryset = queryset.filter(activo=(estado.lower() == 'activo'))
        
        # 4. Ordenamiento
        orden_validos = ['-fecha_registro', 'fecha_registro', 'empleado__primer_nombre', 
                        '-empleado__primer_nombre', 'email', '-email', 'usuario_rol__rol__nombre_rol', 
                        '-usuario_rol__rol__nombre_rol', '-ultimo_login', 'ultimo_login']
        orden = orden if orden in orden_validos else '-fecha_registro'
        queryset = queryset.order_by(orden)
        
        # 5. Paginación
        paginator = Paginator(queryset, 25)
        try:
            usuarios_paginados = paginator.page(page)
        except:
            usuarios_paginados = paginator.page(1)
        
        # 6. Preparar respuesta (versión actualizada para múltiples roles)
        usuarios_data = []
        for user in usuarios_paginados:
            empleado_nombre = 'Sin nombre'
            if hasattr(user, 'empleado') and user.empleado:
                empleado_nombre = f"{user.empleado.primer_nombre or ''} {user.empleado.primer_apellido or ''}".strip()
            
            roles = [ur.rol.nombre_rol for ur in user.usuario_rol_set.all()]
            
            usuarios_data.append({
                'id': user.id,
                'nombre': empleado_nombre if empleado_nombre else 'Sin nombre',
                'email': user.email,
                'roles': roles if roles else ['Sin rol'],
                'activo': user.activo,
                'ultimo_login': user.ultimo_login.strftime('%Y-%m-%d %H:%M:%S') if user.ultimo_login else None,
                'fecha_registro': user.fecha_registro.strftime('%d/%m/%Y %H:%M') if user.fecha_registro else ''
            })
        
        return JsonResponse({
            'success': True,
            'usuarios': usuarios_data,
            'total': paginator.count,
            'paginas': paginator.num_pages,
            'actual': usuarios_paginados.number
        })
        
    except Exception as e:
        logger.error(f"Error en listar_usuarios: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)
    

logger = logging.getLogger(__name__)

@csrf_exempt  # Solo si necesitas desactivar CSRF temporalmente para pruebas
@require_http_methods(["GET", "PUT", "DELETE"])
def manejar_usuario(request, usuario_id):
    """API para ver, editar o eliminar un usuario específico"""
    try:
        # Obtener el usuario
        try:
            user = usuario.objects.select_related('empleado').get(pk=usuario_id)
        except usuario.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuario no encontrado'
            }, status=404)

        if request.method == 'GET':
            # Obtener datos del usuario
            empleado_nombre = 'Sin nombre'
            if hasattr(user, 'empleado') and user.empleado:
                empleado_nombre = f"{user.empleado.primer_nombre or ''} {user.empleado.primer_apellido or ''}".strip()
            
            # Obtener roles del usuario a través de usuario_rol
            roles_qs = usuario_rol.objects.filter(usuario=user).select_related('rol')
            roles_list = []
            for ur in roles_qs:
                rol_obj = ur.rol
                roles_list.append({
                    'id': rol_obj.pk,
                    'codigo': getattr(rol_obj, 'codigo_rol', None),
                    'nombre': getattr(rol_obj, 'nombre_rol', 'Sin rol')
                })
            
            # Formateo seguro de fechas
            def format_date(date):
                if date and isinstance(date, datetime):
                    return date.strftime('%d/%m/%Y %H:%M')
                return None
            
            usuario_data = {
                'id': user.id,
                'nombre': empleado_nombre if empleado_nombre else 'Sin nombre',
                'email': user.email,
                'roles': roles_list,
                'activo': user.activo,
                'ultimo_login': format_date(getattr(user, 'ultimo_login', None)) or 'Nunca',
                'fecha_registro': format_date(getattr(user, 'fecha_registro', None)) or ''
            }
            
            return JsonResponse({
                'success': True,
                'usuario': usuario_data
            })

        elif request.method == 'PUT':
            # Editar usuario
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'error': 'Formato JSON inválido'
                }, status=400)
            
            # Validar datos esenciales
            if 'email' not in data:
                return JsonResponse({
                    'success': False,
                    'error': 'El email es requerido'
                }, status=400)
            
            # Actualizar campos
            user.email = data['email']
            
            if 'roles' in data and data['roles'] is not None:
                try:
                    # roles esperado como lista de ids
                    roles_ids = data['roles']
                    if not isinstance(roles_ids, list):
                        return JsonResponse({
                            'success': False,
                            'error': 'Roles debe ser una lista de IDs'
                        }, status=400)
                    
                    # Validar que todos los roles existan
                    roles_objs = rol.objects.filter(pk__in=roles_ids)
                    if roles_objs.count() != len(roles_ids):
                        return JsonResponse({
                            'success': False,
                            'error': 'Uno o más roles no son válidos'
                        }, status=400)
                    
                    # Eliminar roles antiguos
                    usuario_rol.objects.filter(usuario=user).delete()
                    
                    # Asignar nuevos roles
                    for rol_obj in roles_objs:
                        usuario_rol.objects.create(
                            usuario=user,
                            rol=rol_obj,
                            fecha_asignacion=timezone.now()
                        )
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': f'Error actualizando roles: {str(e)}'
                    }, status=400)
            else:
                # Si no se envían roles, eliminar todos
                usuario_rol.objects.filter(usuario=user).delete()
            
            if 'activo' in data:
                user.activo = bool(data['activo'])
            
            try:
                user.save()
                return JsonResponse({
                    'success': True,
                    'message': 'Usuario actualizado correctamente'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': 'Error al guardar usuario',
                    'detail': str(e)
                }, status=400)

        elif request.method == 'DELETE':
            # Eliminar usuario
            try:
                user.delete()
                return JsonResponse({
                    'success': True,
                    'message': 'Usuario eliminado correctamente'
                })
            except Exception as e:
                logger.error(f"Error eliminando usuario: {str(e)}", exc_info=True)
                return JsonResponse({
                    'success': False,
                    'error': 'Error al eliminar usuario',
                    'detail': str(e)
                }, status=500)

    except Exception as e:
        logger.error(f"Error en manejar_usuario: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)
    
@require_http_methods(["GET"])
def listar_roles(request):
    """API para listar todos los roles disponibles con permisos y cantidad de usuarios (nuevo formato)"""
    try:
        roles = rol.objects.all().order_by('nombre_rol')
        roles_data = []
        
        for r in roles:
            # Obtener permisos asociados al rol
            permisos_qs = permiso.objects.filter(rol_permisos__rol=r)
            permisos_list = list(permisos_qs.values('codigo', 'nombre', 'descripcion'))
            
            # Obtener cantidad de usuarios con este rol (usando la tabla intermedia)
            usuarios_count = usuario_rol.objects.filter(rol_id=r).count()
            
            roles_data.append({
                'codigo_rol': r.codigo_rol,
                'nombre_rol': r.nombre_rol,  # Cambiado a string para compatibilidad con frontend
                'descripcion': r.descripcion,
                'permisos': permisos_list,
                'usuarios_count': usuarios_count,
            })
        
        return JsonResponse({
            'success': True,
            'roles': roles_data,
            'total_roles': len(roles_data)
        })
        
    except Exception as e:
        logger.error(f"Error en listar_roles: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener los roles',
            'detail': str(e)
        }, status=500)

@require_http_methods(["GET"])
def obtener_rol(request, codigo_rol):
    """API para obtener los datos de un rol específico por su código"""
    try:
        rol_obj = rol.objects.get(codigo_rol=codigo_rol)
        permisos_qs = permiso.objects.filter(rol_permisos__rol=rol_obj)
        permisos_list = list(permisos_qs.values('codigo', 'nombre', 'descripcion'))
        usuarios_count = usuario_rol.objects.filter(rol=rol_obj).count()
        rol_data = {
            'codigo_rol': rol_obj.codigo_rol,
            'nombre_rol': rol_obj.nombre_rol,
            'descripcion': rol_obj.descripcion,
            'permisos': permisos_list,
            'usuarios_count': usuarios_count,
        }
        return JsonResponse({
            'success': True,
            'rol': rol_data
        })
    except rol.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Rol no encontrado'
        }, status=404)
    except Exception as e:
        logger.error(f"Error en obtener_rol: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener el rol',
            'detail': str(e)
        }, status=500)

@require_http_methods(["GET"])
def listar_permisos(request):
    """API para listar todos los permisos disponibles"""
    try:
        permisos = permiso.objects.all().order_by('nombre')
        permisos_data = list(permisos.values('codigo', 'nombre', 'descripcion'))
        return JsonResponse({
            'success': True,
            'permisos': permisos_data
        })
    except Exception as e:
        logger.error(f"Error en listar_permisos: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al obtener los permisos',
            'detail': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def crear_roles_api(request):
    """API para crear roles"""
    try:
        # Crear nuevo rol
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        # Validar campos requeridos
        required_fields = ['nombre', 'codigo']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return JsonResponse({
                'success': False,
                'error': f'Campos requeridos: {", ".join(missing_fields)}'
            }, status=400)
        
        # Verificar que el código no exista
        if rol.objects.filter(codigo_rol=data['codigo']).exists():
            return JsonResponse({
                'success': False,
                'error': 'Ya existe un rol con este código'
            }, status=400)
        
        # Crear el rol
        nuevo_rol = rol.objects.create(
            codigo_rol=data['codigo'],
            nombre_rol=data['nombre'],
            descripcion=data.get('descripcion', '')
        )
        
        # Asignar permisos si se proporcionan
        if 'permisos' in data and data['permisos']:
            from .models import rol_permisos
            for permiso_codigo in data['permisos']:
                try:
                    permiso_obj = permiso.objects.get(codigo=permiso_codigo)
                    rol_permisos.objects.create(
                        rol=nuevo_rol,
                        permiso=permiso_obj
                    )
                except permiso.DoesNotExist:
                    continue
        
        return JsonResponse({
            'success': True,
            'message': 'Rol creado exitosamente',
            'rol_id': nuevo_rol.codigo_rol
        })
    
    except Exception as e:
        logger.error(f"Error en crear_roles_api: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["PUT"])
def actualizar_roles(request, rol_id):
    """API para actualizar roles"""
    try:
        # Actualizar rol existente
        if not rol_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de rol requerido'
            }, status=400)
        
        try:
            rol_obj = rol.objects.get(codigo_rol=rol_id)
        except rol.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Rol no encontrado'
            }, status=404)
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Formato JSON inválido'
            }, status=400)
        
        # Actualizar campos
        if 'nombre' in data:
            rol_obj.nombre_rol = data['nombre']
        if 'descripcion' in data:
            rol_obj.descripcion = data['descripcion']
        
        rol_obj.save()
        
        # Actualizar permisos si se proporcionan
        if 'permisos' in data:
            from .models import rol_permisos
            # Eliminar permisos existentes
            rol_permisos.objects.filter(rol=rol_obj).delete()
            
            # Agregar nuevos permisos
            for permiso_codigo in data['permisos']:
                try:
                    permiso_obj = permiso.objects.get(codigo=permiso_codigo)
                    rol_permisos.objects.create(
                        rol=rol_obj,
                        permiso=permiso_obj
                    )
                except permiso.DoesNotExist:
                    continue
        
        return JsonResponse({
            'success': True,
            'message': 'Rol actualizado exitosamente'
        })
    
    except Exception as e:
        logger.error(f"Error en actualizar_roles: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)

@csrf_exempt
@require_http_methods(["DELETE"])
def eliminar_roles(request, rol_id):
    """API para eliminar roles"""
    try:
        # Eliminar rol
        if not rol_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de rol requerido'
            }, status=400)
        
        try:
            rol_obj = rol.objects.get(codigo_rol=rol_id)
        except rol.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Rol no encontrado'
            }, status=404)
        
        # Verificar si hay usuarios asignados a este rol
        usuarios_con_rol = usuario_rol.objects.filter(rol=rol_obj).count()
        if usuarios_con_rol > 0:
            return JsonResponse({
                'success': False,
                'error': f'No se puede eliminar el rol. Hay {usuarios_con_rol} usuario(s) asignado(s) a este rol.'
            }, status=400)
        
        # Eliminar permisos asociados
        from .models import rol_permisos
        rol_permisos.objects.filter(rol=rol_obj).delete()
        
        # Eliminar el rol
        rol_obj.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Rol eliminado exitosamente'
        })
    
    except Exception as e:
        logger.error(f"Error en eliminar_roles: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)


# New API for registro_vacaciones lifecycle management
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime
from django.core.exceptions import ValidationError
from django.db import transaction
from .models import registro_vacaciones, control_vacaciones

#crear registros vacaciones
@csrf_exempt
@require_http_methods(["POST", "PATCH"])
@transaction.atomic
def api_registro_vacaciones(request):
    try:
        if request.method == "POST":
            data = json.loads(request.body)
            control_id = data.get("control_id")
            fecha_inicio = data.get("fecha_inicio")
            fecha_fin = data.get("fecha_fin")
            estado = data.get("estado", "PLAN")  # Valor por defecto 'PLAN' si no se envía

            if not all([control_id, fecha_inicio, fecha_fin]):
                return JsonResponse({"success": False, "message": "Faltan campos requeridos."}, status=400)

            try:
                control = control_vacaciones.objects.get(id=control_id)
            except control_vacaciones.DoesNotExist:
                return JsonResponse({"success": False, "message": "Control de vacaciones no encontrado."}, status=404)

            # Validate dates
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
                fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
                if fecha_fin_dt <= fecha_inicio_dt:
                    return JsonResponse({"success": False, "message": "La fecha de fin debe ser posterior a la fecha de inicio."}, status=400)
            except ValueError:
                return JsonResponse({"success": False, "message": "Formato de fecha inválido."}, status=400)

            registro = registro_vacaciones(
                control=control,
                fecha_inicio=fecha_inicio_dt,
                fecha_fin=fecha_fin_dt,
                estado=estado
            )
            registro.save()

            # Crear asistencias para cada día de vacaciones con estado 'V'
            empleado_obj = control.empleado
            delta_days = (fecha_fin_dt - fecha_inicio_dt).days
            for i in range(delta_days + 1):
                dia_vacacion = fecha_inicio_dt + timedelta(days=i)
                asistencia_obj, created = asistencias.objects.get_or_create(
                    empleado=empleado_obj,
                    fecha=dia_vacacion,
                    defaults={
                        'estado': 'V',
                        'observaciones': 'Vacaciones',
                        'hora_entrada': None,
                        'hora_salida': None
                    }
                )
                if not created:
                    # Actualizar estado y observaciones si ya existe
                    asistencia_obj.estado = 'V'
                    asistencia_obj.observaciones = 'Vacaciones'
                    asistencia_obj.hora_entrada = None
                    asistencia_obj.hora_salida = None
                    asistencia_obj.save()

            return JsonResponse({"success": True, "message": "Vacaciones registradas correctamente.", "id": registro.id})

        elif request.method == "PATCH":
            data = json.loads(request.body)
            registro_id = data.get("id")
            if not registro_id:
                return JsonResponse({"success": False, "message": "ID de registro requerido."}, status=400)

            try:
                registro = registro_vacaciones.objects.get(id=registro_id)
            except registro_vacaciones.DoesNotExist:
                return JsonResponse({"success": False, "message": "Registro de vacaciones no encontrado."}, status=404)

            # Handle inhabilitacion (pausa)
            if data.get("accion") == "inhabilitar":
                fecha_inhabilitacion = data.get("fecha_inhabilitacion")
                motivo = data.get("motivo_inhabilitacion", "")
                if not fecha_inhabilitacion:
                    return JsonResponse({"success": False, "message": "Fecha de inhabilitación requerida."}, status=400)
                try:
                    fecha_inh_dt = datetime.strptime(fecha_inhabilitacion, "%Y-%m-%d").date()
                except ValueError:
                    return JsonResponse({"success": False, "message": "Formato de fecha de inhabilitación inválido."}, status=400)
                try:
                    registro.inhabilitar(fecha_inh_dt, motivo)
                except ValidationError as e:
                    return JsonResponse({"success": False, "message": str(e)}, status=400)
                return JsonResponse({"success": True, "message": "Vacaciones inhabilitadas correctamente."})

            # Handle reanudacion (resume)
            elif data.get("accion") == "reanudar":
                fecha_reanudacion = data.get("fecha_reanudacion")
                if not fecha_reanudacion:
                    return JsonResponse({"success": False, "message": "Fecha de reanudación requerida."}, status=400)
                try:
                    fecha_rean_dt = datetime.strptime(fecha_reanudacion, "%Y-%m-%d").date()
                except ValueError:
                    return JsonResponse({"success": False, "message": "Formato de fecha de reanudación inválido."}, status=400)
                try:
                    registro.reanudar(fecha_rean_dt)
                except ValidationError as e:
                    return JsonResponse({"success": False, "message": str(e)}, status=400)
                return JsonResponse({"success": True, "message": "Vacaciones reanudadas correctamente."})

            else:
                return JsonResponse({"success": False, "message": "Acción no válida."}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)
    


from django.views.decorators.http import require_GET
from login.models import registro_vacaciones

#Funcion listar vacaciones
@csrf_exempt
@require_GET
def listar_registros_vacaciones(request):
    """
    API para listar todos los registros de vacaciones sin filtros.
    """
    try:
        registros = registro_vacaciones.objects.all()

        lista_registros = []
        for reg in registros:
            lista_registros.append({
                'id': reg.id,
                'control_id': reg.control.id if reg.control else None,
                'fecha_inicio': reg.fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': reg.fecha_fin.strftime('%Y-%m-%d'),
                'dias_planificados': reg.dias_planificados,
                'dias_efectivos': reg.dias_efectivos,
                'dias_habilitados': reg.dias_habilitados,
                'estado': reg.estado,
                'fecha_inhabilitacion': reg.fecha_inhabilitacion.strftime('%Y-%m-%d') if reg.fecha_inhabilitacion else None,
                'fecha_reanudacion': reg.fecha_reanudacion.strftime('%Y-%m-%d') if reg.fecha_reanudacion else None,
                'motivo_inhabilitacion': reg.motivo_inhabilitacion,
            })

        return JsonResponse({'success': True, 'registros_vacaciones': lista_registros})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
#Funcion listar permisos_asistencias
@csrf_exempt
@require_http_methods(["GET"])
def listar_permisos_asistencia(request):
    try:
        permisos = permiso_asistencias.objects.select_related('empleado').all()

        lista_registros = []

        for perm in permisos:
            lista_registros.append({
                'id': f'perm_{perm.id}',
                'tipo': f"Permiso ({perm.get_tipo_display()})" if hasattr(perm, 'get_tipo_display') else 'Permiso',
                'empleado': f"{perm.empleado.primer_nombre} {perm.empleado.primer_apellido}" if perm.empleado else '',
                'fecha_inicio': perm.fecha_inicio.strftime('%Y-%m-%d') if hasattr(perm, 'fecha_inicio') and perm.fecha_inicio else '',
                'fecha_fin': perm.fecha_fin.strftime('%Y-%m-%d') if hasattr(perm, 'fecha_fin') and perm.fecha_fin else '',
                'aprobado_por': '',
                'documento_url': '',
            })

        return JsonResponse({'success': True, 'registros': lista_registros})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


logger = logging.getLogger(__name__)
getcontext().prec = 10

def obtener_tasa_bcv():
    try:
        url = "https://www.bcv.org.ve/"
        headers = {
            'User-Agent': 'Mozilla/5.0'
        }
        response = requests.get(url, headers=headers, timeout=15, verify=False)  # ← verify=False
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        tasa_element = soup.find(id='dolar')
        
        if not tasa_element:
            logger.error("Elemento 'dolar' no encontrado")
            return None
            
        tasa_str = tasa_element.find('strong').text.strip()
        tasa_str = tasa_str.replace('.', '').replace(',', '.')
        
        return Decimal(tasa_str)
    except Exception as e:
        logger.error(f"Error BCV: {str(e)}")
        return None

def calcular_cesta_ticket(asistencia_dias, dias_laborables, tasa_bcv=None):
    MONTO_MENSUAL_USD = Decimal('120')
    PORCENTAJE_MINIMO = Decimal('0.75')
    
    if dias_laborables <= 0:
        return Decimal('0')
        
    # Usar siempre el monto mensual completo
    monto_usd = MONTO_MENSUAL_USD
    
    # Solo aplicar descuento por faltas
    asistencia_requerida = Decimal(dias_laborables) * PORCENTAJE_MINIMO
    if Decimal(asistencia_dias) < asistencia_requerida:
        porcentaje_asistencia = Decimal(asistencia_dias) / Decimal(dias_laborables)
        monto_usd *= porcentaje_asistencia
    
    if tasa_bcv and tasa_bcv > Decimal('0'):
        return monto_usd * tasa_bcv
    
    return monto_usd

@csrf_exempt
@transaction.atomic
def generar_nomina_automatica(request):
    if request.method == 'POST':
        try:
            # Obtener tasa BCV al inicio del proceso
            tasa_bcv = obtener_tasa_bcv()
            if tasa_bcv is None:
                logger.warning("No se pudo obtener tasa BCV, usando valor por defecto")
                tasa_bcv = Decimal('36.00')  # Valor por defecto
            
            data = json.loads(request.body)
            logger.info(f"Datos recibidos: {data}")

            # Validar campos obligatorios
            required_fields = ['tipo_nomina', 'mes', 'anio', 'secuencia', 'fecha_cierre', 'conceptos', 'periodo', 'conceptos_individuales']
            if not all(field in data for field in required_fields):
                return JsonResponse({'error': 'Faltan parámetros requeridos'}, status=400)

            # Normalizar secuencia
            secuencia_normalizada = {
                'primera': 'PRIMERA QUINCENA',
                'segunda': 'SEGUNDA QUINCENA',
                'especial': 'ESPECIAL'
            }.get(data['secuencia'].lower(), data['secuencia'])

            # Obtener objetos relacionados
            try:
                tipo_nomina_obj = tipo_nomina.objects.get(tipo_nomina=data['tipo_nomina'])
                mes_obj = meses.objects.get(nombre_mes=data['mes'])
                secuencia_obj = secuencia.objects.get(nombre_secuencia=secuencia_normalizada)
                
                # Obtener TODOS los conceptos seleccionados globalmente (incluyendo deducciones si las hay)
                conceptos_seleccionados_global = concepto_pago.objects.filter(
                    codigo__in=data['conceptos'],
                    status='ACTIVO'
                )
                if conceptos_seleccionados_global.count() != len(data['conceptos']):
                    codigos_invalidos = set(data['conceptos']) - set(conceptos_seleccionados_global.values_list('codigo', flat=True))
                    return JsonResponse({
                        'error': f'Conceptos no encontrados o inactivos: {", ".join(codigos_invalidos)}'
                    }, status=400)
                    
            except tipo_nomina.DoesNotExist:
                return JsonResponse({'error': 'Tipo de nómina no encontrado'}, status=400)
            except meses.DoesNotExist:
                return JsonResponse({'error': 'Mes no encontrado'}, status=400)
            except secuencia.DoesNotExist:
                return JsonResponse({
                    'error': f'Secuencia no válida. Valores aceptados: {list(secuencia_normalizada.values())}'
                }, status=400)
            
            # Crear registro de nómina
            periodo_str = f"{mes_obj.nombre_mes} {data['anio']}"
            nomina_obj = nomina.objects.create(
                tipo_nomina=tipo_nomina_obj,
                periodo=periodo_str,
                secuencia=secuencia_obj,
                fecha_cierre=datetime.strptime(data['fecha_cierre'], '%Y-%m-%d').date(),
                estado='PENDIENTE'  # Estado inicial
            )
            
            # Calcular rango de fechas del periodo
            periodo = int(data['periodo'])
            anio = int(data['anio'])
            mes_num = mes_obj.id_mes if hasattr(mes_obj, 'id_mes') else datetime.strptime(data['mes'], '%B').month
            
            if periodo == 1:
                start_date = datetime(anio, mes_num, 1).date()
                end_date = datetime(anio, mes_num, 15).date()
                dias_laborables = 15
            else:
                start_date = datetime(anio, mes_num, 16).date()
                last_day = monthrange(anio, mes_num)[1]
                end_date = datetime(anio, mes_num, last_day).date()
                dias_laborables = (end_date - start_date).days + 1

            logger.info(f"Periodo: {periodo}, Días: {dias_laborables}, Rango: {start_date} a {end_date}")

            # Función para calcular prima de antigüedad
            def calcular_prima_antiguedad(anios_servicio, salario_asistencias):
                TABLA_PRIMA = {
                    1: 1, 2: 2, 3: 3, 4: 4, 5: 5,
                    6: 6.20, 7: 7.40, 8: 8.60, 9: 9.80,
                    10: 11.00, 11: 12.40, 12: 13.80,
                    13: 15.20, 14: 16.60, 15: 18,
                    16: 19.60, 17: 21.20, 18: 22.80,
                    19: 24.40, 20: 26, 21: 27.80,
                    22: 29.60
                }
                porcentaje = Decimal(str(TABLA_PRIMA.get(anios_servicio, 30)))
                return (salario_asistencias * porcentaje) / Decimal('100')

            # Procesar empleados
            empleados = empleado.objects.filter(status=True)
            stats = {
                'empleados_procesados': 0,
                'conceptos_generados': 0,
                'errores': 0,
                'prenomina_generada': False
            }
            
            # Diccionario para conceptos individuales por empleado: {cedula: [codigo_concepto, ...]}
            conceptos_individuales = data.get('conceptos_individuales', {})

            for emp in empleados:
                detalles_empleado = []  # Lista para almacenar detalles de nómina del empleado
                try:
                    # Validar datos del empleado
                    if not hasattr(emp, 'nivel_salarial') or not emp.nivel_salarial:
                        logger.error(f"Empleado {emp.cedula} sin nivel salarial")
                        stats['errores'] += 1
                        continue
                        
                    if not hasattr(emp, 'fecha_ingreso') or not emp.fecha_ingreso:
                        logger.error(f"Empleado {emp.cedula} sin fecha de ingreso")
                        stats['errores'] += 1
                        continue
                    
                    salario_base = Decimal(str(emp.nivel_salarial.monto))
                    
                    # Calcular asistencia
                    asistencia = asistencias.objects.filter(
                        empleado=emp,
                        fecha__range=(start_date, end_date),
                        estado__in=['F']
                    ).count()

                    asistencia_dias = dias_laborables - asistencia
                    
                    # Calcular salario por asistencias (base para deducciones)
                    salario_diario = salario_base / Decimal('30')
                    salario_asistencias = salario_diario * Decimal(dias_laborables - asistencia)
                    
                    # Calcular años de servicio
                    hoy = date.today()
                    anios_servicio = hoy.year - emp.fecha_ingreso.year
                    if (hoy.month, hoy.day) < (emp.fecha_ingreso.month, emp.fecha_ingreso.day):
                        anios_servicio -= 1
                    
                    # Diccionario para montos calculados
                    montos_conceptos = {}

                    # Obtener conceptos individuales para el empleado si existen, sino usar globales
                    codigos_conceptos_emp_raw = conceptos_individuales.get(str(emp.cedula), data['conceptos'])
                    # Transformar el objeto {concept_code: boolean} a lista de códigos seleccionados
                    if isinstance(codigos_conceptos_emp_raw, dict):
                        codigos_conceptos_emp = [codigo for codigo, seleccionado in codigos_conceptos_emp_raw.items() if seleccionado]
                    else:
                        codigos_conceptos_emp = codigos_conceptos_emp_raw

                    conceptos_seleccionados = concepto_pago.objects.filter(
                        codigo__in=codigos_conceptos_emp,
                        status='ACTIVO'
                    )
                    
                    # Procesar SOLO los conceptos seleccionados para este empleado (incluyendo deducciones si fueron seleccionadas)
                    for concepto in conceptos_seleccionados:
                        monto = Decimal('0')
                        
                        # Conceptos de ingreso
                        if concepto.codigo == '1001':  # Salario por asistencias
                            monto = salario_asistencias
                            montos_conceptos['1001'] = monto
                        
                        elif concepto.codigo == '1101':  # PRM por hijo
                            hijos_activos = hijo.objects.filter(empleado=emp).count()
                            # hijos_activos = hijo.objects.filter(empleado=emp, estudia='S').count()
                            monto = Decimal(hijos_activos) * Decimal('6')
                        
                        elif concepto.codigo == '1102':  # PRM por hijo discapacitado
                            hijos_disc = hijo.objects.filter(empleado=emp, discapacidad=True).count()
                            if hijos_disc > 0:
                                monto = (salario_asistencias * Decimal('0.15')) * Decimal(hijos_disc)
                        
                        elif concepto.codigo == '1103':  # Prima antigüedad
                            if '1001' in montos_conceptos:
                                monto = calcular_prima_antiguedad(anios_servicio, montos_conceptos['1001'])
                        
                        elif concepto.codigo == '1104':  # Prima profesionalización
                            if '1001' in montos_conceptos and hasattr(emp, 'grado_instruccion'):
                                niveles = {
                                    'TSU': Decimal('0.20'),
                                    'PROFESIONAL': Decimal('0.25'),
                                    'ESPECIALISTA': Decimal('0.30'),
                                    'MAESTRIA': Decimal('0.35'),
                                    'DOCTORADO': Decimal('0.40')
                                }
                                porcentaje = niveles.get(emp.grado_instruccion, Decimal('0'))
                                monto = montos_conceptos['1001'] * porcentaje
                                if periodo in [1, 2]:  # Ajuste quincenal
                                    monto /= Decimal('2')
                        
                        elif concepto.codigo == '8003':  # Cesta Ticket
                            monto = calcular_cesta_ticket(asistencia_dias, dias_laborables, tasa_bcv)
                        
                        # Procesar DEDUCCIONES SOLO SI FUERON SELECCIONADAS
                        elif concepto.codigo == '20001':  # IVSS (4% del salario)
                            monto = salario_asistencias * Decimal('0.04') * -1  # Negativo para deducción
                        
                        elif concepto.codigo == '20002':  # Fondo Contributivo RPE (0.5% del salario)
                            monto = salario_asistencias * Decimal('0.005') * -1
                        
                        elif concepto.codigo == '20003':  # FAOV (1% del salario)
                            monto = salario_asistencias * Decimal('0.01') * -1
                        
                        elif concepto.codigo == '20004':  # FPJ (0.5% del salario)
                            monto = salario_asistencias * Decimal('0.005') * -1
                        
                        # Registrar concepto si monto != 0
                        if monto != Decimal('0'):
                            detalle = detalle_nomina.objects.create(
                                nomina=nomina_obj,
                                cedula=emp,
                                codigo=concepto,
                                monto=float(round(monto, 2)))
                            detalles_empleado.append(detalle)
                            stats['conceptos_generados'] += 1
                
                    
                    stats['empleados_procesados'] += 1
                    
                except Exception as e:
                    stats['errores'] += 1
                    logger.error(f"Error procesando empleado {emp.cedula}: {str(e)}", exc_info=True)
            
            # Generar la prenómina después de procesar todos los empleados
            try:
                prenomina_generada = generar_prenomina_para_nomina(nomina_obj)
                if prenomina_generada:
                    stats['prenomina_generada'] = True
                    logger.info(f"Prenómina generada exitosamente para nómina {nomina_obj.id_nomina}")
                else:
                    stats['prenomina_generada'] = False
                    logger.warning(f"No se pudo generar prenómina para nómina {nomina_obj.id_nomina}")
            except Exception as e:
                stats['prenomina_generada'] = False
                logger.error(f"Error al generar prenómina: {str(e)}", exc_info=True)
            
            # Prepare first message summary
            message_summary = "Exitoso"

            return JsonResponse({
                'success': True,
                'nomina_id': nomina_obj.id_nomina,
                'message_summary': message_summary,
                'stats': stats,
                'tasa_bcv_utilizada': float(tasa_bcv),
                'prenomina_generada': stats['prenomina_generada']
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)
        except Exception as e:
            logger.error(f"Error general: {str(e)}", exc_info=True)
            return JsonResponse({'error': f'Error interno: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@csrf_exempt

@transaction.atomic
def aprobar_nomina(request, id_nomina):
    print(f"aprobar_nomina called with id_nomina: {id_nomina}")  # Logging for debugging
    if request.method == 'POST':
        try:
            nomina_obj = nomina.objects.get(id_nomina=id_nomina)
            
            if nomina_obj.estado == 'APROBADA':
                return JsonResponse({'warning': 'Esta nómina ya fue aprobada anteriormente'})
            
            # 1. Cambiar estado
            nomina_obj.estado = 'APROBADA'
            nomina_obj.fecha_aprobacion = timezone.now()
            nomina_obj.save()
            
            # 2. Generar recibos (aquí sí se crean)
            recibos_generados = 0
            empleados_procesados_qs = empleado.objects.filter(
                detalle_nomina__nomina=nomina_obj
            ).distinct()
            empleados_procesados_count = empleados_procesados_qs.count()
            
            for emp in empleados_procesados_qs:
                recibo = recibo_pago.objects.create(
                    nomina=nomina_obj,
                    cedula=emp,
                    fecha_generacion=timezone.now()
                )
                
                # Asociar detalles
                detalles = detalle_nomina.objects.filter(nomina=nomina_obj, cedula=emp)
                for detalle in detalles:
                    detalle_recibo.objects.create(
                        recibo=recibo,
                        detalle_nomina=detalle
                    )
                
                recibos_generados += 1
            
            return JsonResponse({
                'success': True,
                'message': 'Nómina aprobada correctamente',
                'stats': {
                    'recibos_generados': recibos_generados,
                    'empleados_procesados': empleados_procesados_count
                },
                'fecha_aprobacion': nomina_obj.fecha_aprobacion.strftime('%Y-%m-%d %H:%M')
            })
            
        except nomina.DoesNotExist:
            return JsonResponse({'error': 'Nómina no encontrada'}, status=404)
        except Exception as e:
            logger.error(f"Error al aprobar: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@require_http_methods(["GET"])
def api_empleados_por_tipo(request, tipo):
    """
    API para obtener empleados filtrados por tipo de nómina
    """
    try:
        logger.info(f"api_empleados_por_tipo called with tipo: {tipo}")
        # Obtener el tipo de trabajador por descripción

        try:
            tipo_trabajador_obj = tipo_trabajador.objects.filter(descripcion__icontains=tipo).first()
        except tipo_trabajador.DoesNotExist:
            logger.warning(f"Tipo de trabajador no encontrado: {tipo}")
            return JsonResponse({
                'success': False,
                'error': f'Tipo de trabajador "{tipo}" no encontrado'
            }, status=404)
        
        # Filtrar empleados activos por tipo de trabajador
        empleados_qs = empleado.objects.filter(
            tipo_trabajador=tipo_trabajador_obj,
            status=True
        ).select_related('cargo', 'cargo__familia', 'cargo__nivel', 'tipo_trabajador').order_by('primer_apellido', 'primer_nombre')
        
        empleados_data = []
        for emp in empleados_qs:
            try:
                cargo_nombre = emp.cargo.nombre_completo if emp.cargo else 'Sin cargo'
                tipo_trab_desc = emp.tipo_trabajador.descripcion if emp.tipo_trabajador else 'Desconocido'
                empleados_data.append({
                    'cedula': emp.cedula,
                    'nombre_completo': emp.get_nombre_completo(),
                    'cargo': cargo_nombre,
                    'tipo_trabajador': tipo_trab_desc
                })
            except Exception as inner_e:
                logger.error(f"Error procesando empleado {emp.cedula}: {str(inner_e)}", exc_info=True)
                # Skip this employee or add partial data
                continue
        
        return JsonResponse({
            'success': True,
            'empleados': empleados_data,
            'total': len(empleados_data)
        })
        
    except Exception as e:
        logger.error(f"Error en api_empleados_por_tipo: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)

@require_http_methods(["GET"])
def api_conceptos(request):
    """
    API para obtener todos los conceptos de pago disponibles
    """
    try:
        # Obtener conceptos activos ordenados por código
        conceptos_qs = concepto_pago.objects.filter(
            status='ACTIVO'
        ).select_related('tipo_pago').order_by('codigo')
        
        conceptos_data = []
        for concepto in conceptos_qs:
            conceptos_data.append({
                'codigo': concepto.codigo,
                'descripcion': concepto.descripcion,
                'tipo_concepto': concepto.tipo_concepto,
                'tipo_pago': concepto.tipo_pago.nombre_tipopago if concepto.tipo_pago else '',
                'nombre_nomina': concepto.nombre_nomina,
                'status': concepto.status
            })
        
        return JsonResponse({
            'success': True,
            'conceptos': conceptos_data,
            'total': len(conceptos_data)
        })
        
    except Exception as e:
        logger.error(f"Error en api_conceptos: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Error al procesar la solicitud',
            'detail': str(e)
        }, status=500)


@require_GET
def generar_arc(request):
    import calendar
    from decimal import Decimal
    from django.db.models import Sum
    try:
        cedula = request.GET.get('cedula')
        anio = int(request.GET.get('anio'))
        
        empleado_obj = empleado.objects.get(cedula=cedula)
        
        # Verificar si ya existe un ARC para este año
        arc, created = ARC.objects.get_or_create(
            empleado=empleado_obj,
            anio=anio,
            defaults={
                'total_monto_declarar': 0,
                'islr_total_retenido': 0
            }
        )
        
        def crear_detalles_arc(arc_instance):
            # Eliminar detalles existentes para evitar duplicados
            arc_instance.detalles.all().delete()
            
            meses = [(i, calendar.month_name[i]) for i in range(1, 13)]
            
            total_monto_declarar = Decimal('0')
            total_islr_retenido = Decimal('0')
            
            for mes_num, nombre_mes in meses:
                # Aquí se puede obtener datos reales de recibo_pago o cálculos específicos
                # Por simplicidad, se asignan valores de ejemplo o cero
                monto_bruto = Decimal('0')
                porcentaje_retencion = Decimal('0')
                islr_retenido = Decimal('0')
                monto_declarar = Decimal('0')
                
                # Ejemplo: sumar montos de recibo_pago para el empleado, año y mes
                recibos_mes = recibo_pago.objects.filter(
                    cedula=arc_instance.empleado,
                    fecha_generacion__year=arc_instance.anio,
                    fecha_generacion__month=mes_num
                )
                monto_bruto_mes = recibos_mes.aggregate(total=Sum('detalles__detalle_nomina__monto'))['total'] or Decimal('0')
                monto_bruto = monto_bruto_mes
                
                # Ejemplo: calcular islr_retenido y porcentaje_retencion (puede ajustarse según reglas)
                # Aquí se asigna 0 para simplificar, se puede mejorar con lógica real
                porcentaje_retencion = Decimal('0')
                islr_retenido = Decimal('0')
                monto_declarar = monto_bruto  # Asumir monto declarar igual a bruto para ejemplo
                
                total_monto_declarar += monto_declarar
                total_islr_retenido += islr_retenido
                
                DetalleARC.objects.create(
                    arc=arc_instance,
                    mes=mes_num,
                    nombre_mes=nombre_mes,
                    especificacion_superior='OTROS',
                    especificacion_inferior='',
                    monto_bruto=monto_bruto,
                    porcentaje_retencion=porcentaje_retencion,
                    islr_retenido=islr_retenido,
                    monto_declarar=monto_declarar
                )
            
            arc_instance.total_monto_declarar = total_monto_declarar
            arc_instance.islr_total_retenido = total_islr_retenido
            arc_instance.save()
        
        if created:
            crear_detalles_arc(arc)
        else:
            # Si ya existía, recalcular los totales
            arc.total_monto_declarar = sum(d.monto_declarar for d in arc.detalles.all())
            arc.islr_total_retenido = sum(d.islr_retenido for d in arc.detalles.all())
            arc.save()
        
        return JsonResponse({
            'success': True,
            'arc_id': arc.id_arc
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@require_GET
def datos_arc(request, arc_id):
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"datos_arc called with arc_id={arc_id}")
    try:
        arc = ARC.objects.get(id_arc=arc_id)
        detalles = arc.detalles.all().order_by('mes')
        logger.info(f"datos_arc: detalles count = {detalles.count()} for arc_id={arc_id}")

        from django.db.models import Sum, F, Q
        from django.db.models.functions import ExtractMonth

        empleado_obj = arc.empleado
        anio = arc.anio

        recibos = (
            recibo_pago.objects.filter(
                cedula=empleado_obj,
                fecha_generacion__year=anio
            )
            .annotate(mes=ExtractMonth('fecha_generacion'))
            .values('mes')
            .annotate(
                suma_monto=Sum(
                    'detalles__detalle_nomina__monto',
                    filter=Q(detalles__detalle_nomina__codigo__codigo='1001')
                )
            )
            .order_by('mes')
        )
        recibos_count = recibos.count()
        logger.info(f"datos_arc: recibos count = {recibos_count} for empleado={empleado_obj} year={anio}")

        monto_por_mes = {item['mes']: item['suma_monto'] or 0 for item in recibos}

        resumen = {}
        for d in detalles:
            desc = d.especificacion_superior or 'OTROS'
            if desc not in resumen:
                resumen[desc] = 0
            resumen[desc] += float(d.monto_bruto)

        resumen['ISLR RETENIDO'] = float(arc.islr_total_retenido)

        meses_actualizados = []
        for d in detalles:
            mes_num = d.mes
            suma_monto = monto_por_mes.get(mes_num, 0)
            meses_actualizados.append({
                'nombre_mes': d.nombre_mes,
                'especificacion_superior': d.especificacion_superior,
                'especificacion_inferior': d.especificacion_inferior,
                'monto_bruto': float(suma_monto),
                'porcentaje_retencion': float(d.porcentaje_retencion),
                'islr_retenido': float(d.islr_retenido),
                'monto_declarar': float(suma_monto)
            })

        logger.info(f"datos_arc: meses_actualizados length = {len(meses_actualizados)}")

        return JsonResponse({
            'success': True,
            'data': {
                'anio': arc.anio,
                'agente': {
                    'nombre': 'POLICÍA DEL ESTADO ANZOÁTEGUI',
                    'rif': 'G-20001091-6',
                    'direccion': 'Av. Intercomunal Jorge Rodríguez, Lechería'
                },
                'usuario': {
                    'nombre_completo': arc.empleado.get_nombre_completo(),
                    'cedula': arc.empleado.cedula
                },
                'meses': meses_actualizados,
                'total_monto_declarar': float(arc.total_monto_declarar),
                'resumen': resumen,
                'nota': 'Este comprobante es válido para la declaración del ISLR del año correspondiente.'
            }
        })

    except ARC.DoesNotExist:
        logger.error(f"datos_arc: ARC with id {arc_id} does not exist")
        return JsonResponse({
            'success': False,
            'error': 'ARC no encontrado'
        }, status=404)
    except Exception as e:
        logger.error(f"datos_arc: Exception {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@csrf_exempt
@require_GET
def listar_arc(request):
    cedula = request.GET.get('cedula')
    if not cedula:
        return JsonResponse({'success': False, 'error': 'Cédula no proporcionada'}, status=400)
    try:
        empleado_obj = empleado.objects.get(cedula=cedula)
        arcs = ARC.objects.filter(empleado=empleado_obj).order_by('-anio')
        arcs_data = []
        for arc in arcs:
            arcs_data.append({
                'id_arc': arc.id_arc,
                'anio': arc.anio,
                'total_monto_declarar': float(arc.total_monto_declarar),
                'islr_total_retenido': float(arc.islr_total_retenido)
            })
        return JsonResponse({'success': True, 'arcs': arcs_data})
    except empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Empleado no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


